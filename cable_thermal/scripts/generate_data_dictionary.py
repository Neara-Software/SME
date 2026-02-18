"""Generate a CYMCAP Database Data Dictionary Excel workbook.

Connects to the sample CymcapData.MDB, reads schema metadata,
merges with a comprehensive knowledge dictionary, and outputs
one Excel sheet per table with column descriptions.

Usage:
    uv run --with openpyxl python scripts/generate_data_dictionary.py
"""

from __future__ import annotations

import sys
from pathlib import Path

import pyodbc
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_DIR = SCRIPT_DIR.parent
DB_PATH = PROJECT_DIR / "data" / "sample" / "CymcapData.MDB"
OUTPUT_PATH = PROJECT_DIR / "data" / "sample" / "CYMCAP_DataDictionary.xlsx"

# ---------------------------------------------------------------------------
# Enum value strings (from enums.py / CYMCAP_CABLES.xlsx analysis)
# ---------------------------------------------------------------------------
INSTALL_ENUM = "1=Pipe (Triangular), 2=LPOF, 3=Concentric Neutral, 4=Extruded, 5=Other, 6=Combined Sheath"
CONDMAT_ENUM = "1=Copper, 2=Aluminium"
INSTYP_ENUM = "-1=Custom, 1=Solid, 2=LPOF, 3=HPOF Self-Contained, 4=HPOF Pipe, 9=EPR, 10=PVC, 11=Polyethylene, 12=XLPE (unfilled), 13=XLPE (filled), 14=Paper"
DIELEC_ENUM = "-3=Enter tan delta & epsilon, -1=Enter tan delta, 0=Program selects"
SKIDCON_ENUM = "-1=Custom, 0=None, 1=Copper, 2=Brass/Bronze, 3=Zinc, 4=Stainless Steel"
SHEAMAT_ENUM = "0=None, 1=Aluminium, 2=Lead, 3=Lead w/ Reinforcing, 4=Copper"
CORRUG_ENUM = "1=Non-corrugated, 2=Corrugated, 3=Longitudinally Applied Tape"
JACKMAT_ENUM = "-1=Custom, 0=None, 1=Compound Jute, 2=Rubber Sandwich, 3=Polychloroprene, 4=Polyethylene, 5=PVC, 6=EPR, 7=Butyl Rubber, 8=Coal Tar Rubber"
ARMREIN_ENUM = "-4=Custom non-mag tape, -3=Custom mag wire/tape, -1=Custom non-mag wire, 0=None, 1=Steel wire touching, 2=Steel wire not touching, 3=Steel tape, 4=Copper wire, 5=Stainless steel wire, 6=TECK"
ARMBED_ENUM = "-1=Custom, 0=None, 1=Compounded Jute, 2=Rubber Sandwich"
ARMSERV_ENUM = "-1=Custom, 0=None, 1=Compounded Jute"
INSUSHI_ENUM = "0=No Screen, 1=Belted (unscreened), 2=Semiconductive, 3=Copper Tape, 4=Aluminium Tape"
TAPINPIP_ENUM = "0=None, 2=Brass/Bronze, 4=Stainless Steel, 5=Steel"
ACDC_ENUM = "0=AC, 1=DC"
CONDSHLDMAT_ENUM = "-1=Custom, 0=Same as insulation, 15=Semiconductive, 99=Undefined"
SKNROUNDWI_ENUM = "1=Round Wires, 2=Flat Straps"
NUMBCON_ENUM = "1=Single Core, 3=Three Core"
CONDCON_ENUM = "1-11: conductor construction code (round stranded, solid, segmental, etc.)"
PIPCONFI_ENUM = "0=N/A, 1=Triangular, 2=Cradled"
FILLERS_ENUM = "-99=No Fillers, 0=Custom, 1=Same as insulation"
MODECALC_ENUM = "Calculation mode code"
INSTTYPE_ENUM = "1=Underground, 2=Submarine, etc."

# ---------------------------------------------------------------------------
# Knowledge Dictionary
# ---------------------------------------------------------------------------
# Key: (TABLE_NAME, COLUMN_NAME) -> dict with units, description, enum_values, justification
# Convention: TABLE_NAME and COLUMN_NAME are UPPER CASE for matching

def _build_knowledge_dict() -> dict[tuple[str, str], dict]:
    """Build the comprehensive knowledge dictionary."""
    k: dict[tuple[str, str], dict] = {}

    def add(table: str, col: str, units: str, desc: str, enum: str = "", just: str = ""):
        k[(table.upper(), col.upper())] = {
            "units": units,
            "description": desc,
            "enum_values": enum,
            "justification": just,
        }

    # ===================================================================
    # CABLES (231 columns)
    # ===================================================================
    T = "CABLES"

    # Identity
    add(T, "IDCAB", "", "Cable ID — unique identifier within CYMCAP", "", "Primary key, VARCHAR")
    add(T, "CABTITLE", "", "Cable title / description text", "", "User-entered label")
    add(T, "COMMENTS", "", "Free-text comments", "", "User-entered notes")
    add(T, "TAG", "", "Record tag (typically 'F')", "", "Internal flag; observed value 'F'")

    # Cable type
    add(T, "INSTALL", "", "Installation / cable type code", INSTALL_ENUM, "Maps to InstallationType enum per CYMCAP UI")
    add(T, "INSTSPEC", "", "Installation specification (currently unused)", "", "All values blank in sample DB")
    add(T, "CONDMAT", "", "Conductor material code", CONDMAT_ENUM, "Maps to ConductorMaterial enum")
    add(T, "CONDCON", "", "Conductor construction code", CONDCON_ENUM, "Construction type (round stranded, solid, segmental, etc.)")
    add(T, "DRYIMPR", "", "Dry/impregnated insulation flag", "0,1,2,99", "Insulation paper treatment option")
    add(T, "CONLOS", "", "Conductor loss calculation option", "-1,0,1", "Controls how conductor losses are computed")
    add(T, "INSTYP", "", "Insulation type code", INSTYP_ENUM, "Maps to InsulationType enum")
    add(T, "DIELEC", "", "Dielectric loss option", DIELEC_ENUM, "Maps to DielectricOption enum")
    add(T, "NUMCON", "", "Number of conductors (legacy, appears unused)", NUMBCON_ENUM, "See NUMBCON instead")
    add(T, "SKIDCON", "", "Concentric/skid wire material code", SKIDCON_ENUM, "Maps to ConcentricNeutralMaterial enum")
    add(T, "SKNROUNDWI", "", "Concentric wire construction type", SKNROUNDWI_ENUM, "Round wires vs flat straps")
    add(T, "SHEAMAT", "", "Sheath material code", SHEAMAT_ENUM, "Maps to SheathMaterial enum")
    add(T, "CORRUG", "", "Corrugation type", CORRUG_ENUM, "Maps to CorrugationType enum")
    add(T, "BONDING", "", "Bonding configuration (currently unused)", "", "All values blank in sample DB")
    add(T, "LOSFACT", "", "Loss factor override (currently unused)", "", "All values blank in sample DB")
    add(T, "JACKMAT", "", "Jacket material code", JACKMAT_ENUM, "Maps to JacketMaterial enum")
    add(T, "ARMREIN", "", "Armour type code", ARMREIN_ENUM, "Maps to ArmourType enum")
    add(T, "ARMBED", "", "Armour bedding material code", ARMBED_ENUM, "Maps to ArmourBedding enum")
    add(T, "PERMEAB", "", "Relative permeability of armour", "", "Typically 0 (program default); for magnetic armour")
    add(T, "ARMSERV", "", "Armour serving material code", ARMSERV_ENUM, "Maps to ArmourServing enum")
    add(T, "INSUSHI", "", "Insulation screen type", INSUSHI_ENUM, "Maps to InsulationScreenType enum")
    add(T, "TAPINPIP", "", "Reinforcing tape over insulation screen material", TAPINPIP_ENUM, "Material code for sheath reinforcing tape")
    add(T, "OVERSHEATH", "", "Oversheath flag (currently unused)", "", "All zeros in sample DB")
    add(T, "TRANSPO", "", "Transposition flag (currently unused)", "", "All blank in sample DB")
    add(T, "PIPCONFI", "", "Pipe configuration", PIPCONFI_ENUM, "Triangular vs cradled for pipe-type cables")
    add(T, "PIPEMAT", "", "Pipe material code (currently unused)", "", "Not populated in sample DB")
    add(T, "DUCTBANK", "", "Duct bank reference (currently unused)", "", "Not populated in sample DB")
    add(T, "MINSECLE", "", "Minimum section length (currently unused)", "", "Not populated in sample DB")
    add(T, "SECSPAC", "", "Section spacing (currently unused)", "", "Not populated in sample DB")
    add(T, "NSEGMENTS", "", "Number of segments for FEM meshing", "1,6", "Conductor mesh subdivision")
    add(T, "BRATIO", "", "Beta ratio (bundle conductor ratio)", "", "Typically 1.0")
    add(T, "RHC", "K.m/W", "Custom conductor thermal resistivity", "", "Zero = program default")
    add(T, "ALFAC", "1/K", "Conductor resistance temperature coefficient", "", "Zero = program default")
    add(T, "AKP", "", "Proximity effect factor kp (IEC 60287-1-1)", "", "0 = program selects; manual override for AC resistance")
    add(T, "AKS", "", "Skin effect factor ks (IEC 60287-1-1)", "", "0 = program selects; 0.001 stored when GUI shows 0.0")
    add(T, "RHI", "K.m/W", "Custom insulation thermal resistivity", "", "Used when INSTYP=-1 (custom)")
    add(T, "COSPHI", "", "Insulation dielectric loss tangent (tan delta)", "", "Custom value used when DIELEC=-1 or -3")
    add(T, "EPSILN", "", "Insulation relative permittivity (epsilon_r)", "", "Custom value used when DIELEC=-3")
    add(T, "RHK", "ohm.m", "Custom concentric neutral wire electrical resistivity", "", "Used when SKIDCON=-1 (custom)")
    add(T, "ALFAK", "1/K", "Concentric neutral resistance temperature coefficient", "", "Per K at 20°C")
    add(T, "RHS", "ohm.m", "Custom sheath electrical resistivity", "", "Zero = program default for selected material")
    add(T, "ALFAS", "1/K", "Sheath resistance temperature coefficient", "", "Zero = program default")
    add(T, "ALOS", "", "Additional losses factor (currently unused)", "", "All blank in sample DB")
    add(T, "RHJ", "K.m/W", "Custom jacket thermal resistivity", "", "Used when JACKMAT=-1 (custom)")
    add(T, "RHA", "ohm.m", "Custom armour electrical resistivity at 20°C", "", "Value in Ω·m (e.g. 1.38e-7)")
    add(T, "ALFAA", "1/K", "Armour resistance temperature coefficient at 20°C", "", "Per K")
    add(T, "RHB", "K.m/W", "Armour bedding thermal resistivity", "", "Used when ARMBED=-1 (custom)")
    add(T, "AME", "", "Unknown material parameter (always 0)", "", "Currently zero for everything")
    add(T, "AMT", "", "Unknown material parameter (always 0)", "", "Currently zero for everything")
    add(T, "GAMMA", "", "Unknown material parameter (always 0)", "", "Currently zero for everything")
    add(T, "RHAS", "K.m/W", "Armour serving thermal resistivity", "", "Used when ARMSERV=-1 (custom)")
    add(T, "RHT", "K.m/W", "Unknown thermal resistivity (always 0)", "", "Currently zero for everything")
    add(T, "ALFAT", "1/K", "Unknown temperature coefficient (always 0)", "", "Currently zero for everything")
    add(T, "RHO", "", "Unknown resistivity parameter (always 0)", "", "Currently zero for everything")
    add(T, "PIPFAC", "", "Pipe factor (currently unused)", "", "All blank in sample DB")
    add(T, "AM", "", "Unknown parameter (currently unused)", "", "All blank in sample DB")
    add(T, "AN", "", "Unknown parameter (currently unused)", "", "All blank in sample DB")
    add(T, "NSECT", "", "Number of sections (currently unused)", "", "All blank in sample DB")
    add(T, "RHD", "", "Unknown thermal resistivity (currently unused)", "", "All blank in sample DB")

    # Physical dimensions
    add(T, "NUMBCON", "", "Number of cores/conductors", NUMBCON_ENUM, "1=Single Core, 3=Three Core")
    add(T, "CONAREA", "m²", "Conductor cross-sectional area", "", "Standard cable conductor area in m²")
    add(T, "CONDIAM", "m", "Conductor outer diameter", "", "Diameter in metres")
    add(T, "CONSHDI", "m", "Diameter over conductor shield", "", "Outer diameter after conductor shield layer")
    add(T, "CONSPAC", "m", "Three-core cable conductor spacing diameter", "", "Diameter after individual cores in 3-core cable")
    add(T, "CONINDI", "m", "Conductor inner diameter (hollow core)", "", "For hollow-core/oil-channel conductors")
    add(T, "CONEQDI", "m", "Conductor equivalent diameter", "", "Used for sector-shaped conductors")
    add(T, "CENTRDIS", "m", "Centre distance between conductors", "", "Only populated for specific cable types")
    add(T, "RADCIRC", "m", "Radius of circle through 3-core conductors", "", "For 3-core sector-shaped cables")
    add(T, "VOLTAGE", "V", "Rated voltage", "", "Phase-to-ground voltage in Volts")
    add(T, "INSUDIA", "m", "Insulation outer diameter", "", "Diameter over insulation in metres")
    add(T, "INSHIDI", "m", "Insulation inner (shield) diameter", "", "Inner diameter of insulation screen layer")
    add(T, "INSTHIC", "m", "Insulation thickness", "", "For 3-core: 2× insulation thickness")
    add(T, "CINSSHTH", "m", "Common insulation screen thickness (3-core)", "", "2× insulation screen thickness for 3-core cables")
    add(T, "INSSHTH", "m", "Insulation screen thickness (individual)", "", "Per-core insulation screen thickness")
    add(T, "MESHITH", "m", "Metallic insulation screen thickness", "", "Thickness of metallic insulation screen")
    add(T, "CIRCIRDI", "m", "Unknown diameter (rare usage)", "", "Only populated in one cable (case 51)")
    add(T, "SHEATHDI", "m", "Sheath outer diameter", "", "Metallic sheath outer diameter in metres")
    add(T, "SHETHIC", "m", "Sheath thickness", "", "Metallic sheath thickness in metres")
    add(T, "SHINRAD", "m", "Corrugated sheath inner radius", "", "Inner trough radius for corrugated sheaths")
    add(T, "SHOUTRAD", "m", "Corrugated sheath outer radius", "", "Outer crest radius for corrugated sheaths")
    add(T, "TAPBIND", "m", "Tape over insulation screen — outer diameter", "", "Diameter after tape application")
    add(T, "TAPBINTH", "m", "Tape over insulation screen — thickness", "", "Tape thickness in metres")
    add(T, "TAPBINWI", "m", "Tape over insulation screen — width", "", "Tape width in metres")
    add(T, "TAPENUM", "", "Number of tapes over insulation screen", "", "Integer count")
    add(T, "LAYLENG", "m", "Tape lay length", "", "Helical lay length of tape")
    add(T, "LAYLENG1", "m", "Tape lay length (secondary, unclear usage)", "", "Purpose unclear")
    add(T, "OVSHETDI", "m", "Armour bedding outer diameter", "", "Diameter after armour bedding; rarely populated")
    add(T, "OVSHETH", "m", "Armour bedding thickness (alternative)", "", "Rarely populated")
    add(T, "JACKDIA", "m", "Jacket outer diameter", "", "Non-metallic jacket outer diameter")
    add(T, "JACKTHI", "m", "Jacket thickness", "", "Non-metallic jacket thickness")
    add(T, "ARMBEDTH", "m", "Armour bedding thickness", "", "Bedding layer thickness in metres")
    add(T, "ARMSERTH", "m", "Armour serving thickness", "", "Serving layer thickness in metres")
    add(T, "SKINEUDI", "m", "Diameter over concentric neutral wires", "", "Outer diameter after concentric wire layer")
    add(T, "SKNSTRAPLG", "m", "Concentric neutral strap width", "", "Width of flat strap (when flat straps selected)")
    add(T, "WIARMDIA", "m", "Diameter over wire armour", "", "Outer diameter after armour wire layer")
    add(T, "TAPARMDI", "m", "Diameter over tape armour", "", "Outer diameter after armour tape layer")
    add(T, "SKNEWIDI", "m", "Concentric neutral wire diameter", "", "Individual wire diameter")
    add(T, "ARMWIRDI", "m", "Armour wire diameter", "", "Individual armour wire diameter")
    add(T, "TAPARMAR", "m²", "Armour tape cross-sectional area", "", "For tape armour where width not separately defined")
    add(T, "SKNEWILE", "m", "Concentric neutral wire lay length", "", "Helical lay length")
    add(T, "ARMWIRLE", "m", "Armour wire lay length", "", "Helical lay length")
    add(T, "SKNEWINU", "", "Number of concentric neutral wires", "", "Integer count")
    add(T, "ARMWIRNU", "", "Number of armour wires", "", "Integer count")
    add(T, "OVERDIAM", "m", "Overall cable outer diameter", "", "Final outer diameter of complete cable")
    add(T, "DUPIINDI", "m", "Duct/pipe inner diameter (currently unused)", "", "Not populated in sample DB")
    add(T, "DUPIOTDI", "m", "Duct/pipe outer diameter (currently unused)", "", "Not populated in sample DB")
    add(T, "SLCSHTH", "", "Sheath configuration for 3-core cables", "1=Individual, 2=Common", "Whether sheath is per-core or common")
    add(T, "SLCCONC", "", "Concentric wire configuration for 3-core cables", "1=Individual, 2=Common", "Whether concentric wires are per-core or common")
    add(T, "CORRUGHGT", "m", "Corrugation height", "", "Height of corrugation on corrugated sheath")
    add(T, "TAPOVERLAP", "m", "Tape overlap", "", "Overlap distance for helically applied tapes")
    add(T, "TAPWIDTH", "m", "Tape width", "", "Width of reinforcing/screen tape")
    add(T, "TAPTH", "m", "Tape thickness", "", "Thickness of reinforcing/screen tape")
    add(T, "ACCFACTOR", "", "Accumulation factor", "", "Purpose unclear; possibly for bundled cables")
    add(T, "DCRESIS", "ohm/m", "DC resistance at 20°C", "", "Conductor DC resistance per metre")
    add(T, "DOVLCSHLD", "m", "Diameter over longitudinal cable shield", "", "Only for longitudinally applied tapes")
    add(T, "LOCKEDBY", "", "Locked by user (currently unused)", "", "Multi-user locking field")
    add(T, "LOCKTSTAMP", "", "Lock timestamp (currently unused)", "", "Multi-user locking timestamp")

    # Custom material names
    add(T, "CONDCNAM", "", "Custom conductor material name", "", "Text label for custom conductor")
    add(T, "INSUCNAM", "", "Custom insulation material name", "", "Text label for custom insulation")
    add(T, "SHTHCNAM", "", "Custom sheath material name", "", "Text label for custom sheath")
    add(T, "REINFCNAM", "", "Custom reinforcing tape material name", "", "Text label for custom reinforcing")
    add(T, "SKIDCNAM", "", "Custom concentric/skid wire material name", "", "Text label for custom wire material")
    add(T, "OVSHTHCNAM", "", "Custom oversheath material name", "", "Text label")
    add(T, "ARMBEDCNAM", "", "Custom armour bedding material name", "", "Text label")
    add(T, "ARMOURCNAM", "", "Custom armour material name", "", "Text label")
    add(T, "ARMSRVCNAM", "", "Custom armour serving material name", "", "Text label")
    add(T, "JACKETCNAM", "", "Custom jacket material name", "", "Text label")

    # Temperatures
    add(T, "SS_TEMP", "°C", "Maximum steady-state conductor temperature", "", "IEC 60287 maximum operating temperature")
    add(T, "TR_TEMP", "°C", "Maximum transient/emergency conductor temperature", "", "IEC 60853 emergency rating temperature")
    add(T, "MSTR_IDCAB", "", "Master cable ID reference (currently unused)", "", "Not populated in sample DB")

    # Short circuit data fields
    add(T, "ShCctConductorData", "", "Short-circuit conductor data (CSV string)", "", "Format: rho,alpha,maxT,duration,flag")
    add(T, "ShCctSheathData", "", "Short-circuit sheath data (CSV string)", "", "Format: rho,alpha,maxT,duration,flag")
    add(T, "ShCctReinforcingTapeData", "", "Short-circuit reinforcing tape data (CSV string)", "", "Format: rho,alpha,maxT,duration,flag")
    add(T, "ShCctSkidConcentricData", "", "Short-circuit concentric/skid data (CSV string)", "", "Format: rho,alpha,maxT,duration,flag")
    add(T, "ShCctArmourData", "", "Short-circuit armour data (CSV string)", "", "Format: rho,alpha,maxT,duration,flag")
    add(T, "ShCctPipeCoatingMaterial", "", "Short-circuit pipe coating material", "", "Not populated")
    add(T, "ShCctPipeCoatingCustomRHJ", "K.m/W", "Custom pipe coating thermal resistivity", "", "Not populated")
    add(T, "ShCctPipeMaterial", "", "Short-circuit pipe material code", "", "Not populated")
    add(T, "ShCctPipeFactor", "", "Short-circuit pipe factor", "", "Not populated")
    add(T, "ShCctInsideDuctDiam", "m", "Short-circuit inside duct diameter", "", "Not populated")
    add(T, "ShCctOutsideDuctDiam", "m", "Short-circuit outside duct diameter", "", "Not populated")
    add(T, "ShCctOverallDuctDiam", "m", "Short-circuit overall duct diameter", "", "Not populated")

    # Fillers and specialty fields
    add(T, "FillersSameAsInsuFlag", "", "Filler material same-as-insulation flag", FILLERS_ENUM, "For 3-core cable fillers")
    add(T, "FillersResis", "K.m/W", "Filler thermal resistivity", "", "Custom filler thermal resistivity")
    add(T, "NbrOfWiresComposingStrandedConductor", "", "Number of strands in conductor", "", "Used for inductance calculations")
    add(T, "WireGauge", "", "Wire gauge selection code", "", "Concentric wire gauge; affects thickness/diameter")
    add(T, "CasingInsideDiam", "m", "Casing inside diameter (unused)", "", "Not populated")
    add(T, "CasingOutsideDiam", "m", "Casing outside diameter (unused)", "", "Not populated")
    add(T, "CasingConst", "", "Casing construction code (unused)", "", "Not populated")
    add(T, "CasingRh", "K.m/W", "Casing thermal resistivity (unused)", "", "Not populated")
    add(T, "CasingFormation", "", "Casing formation code (unused)", "", "Not populated")
    add(T, "ArmourEmbeddedPEInnerTh", "m", "PE/AL/PE armour — inner PE thickness", "", "For PE/AL/PE laminate armour")
    add(T, "ArmourEmbeddedAluminumTh", "m", "PE/AL/PE armour — aluminium thickness", "", "For PE/AL/PE laminate armour")
    add(T, "ArmourEmbeddedPEOuterTh", "m", "PE/AL/PE armour — outer PE thickness", "", "For PE/AL/PE laminate armour")
    add(T, "PrintingCompactCompressedFlag", "", "Compact vs compressed conductor label flag", "0=Compressed, 1=Compact", "Display label only")
    add(T, "SideBySidePictFileName", "", "Side-by-side picture filename (GUI)", "", "GUI display field")
    add(T, "SideBySidePictScaleIndex", "", "Picture scale index (GUI)", "", "GUI display field")
    add(T, "DataSheetFileNames", "", "Attached data sheet filenames (GUI)", "", "GUI display field")
    add(T, "GMRConductor", "m", "Geometric mean radius — conductor", "", "For impedance calculations")
    add(T, "GMRSheath", "m", "Geometric mean radius — sheath", "", "For impedance calculations")
    add(T, "GMRConcentricNeutral", "m", "Geometric mean radius — concentric neutral", "", "For impedance calculations")
    add(T, "LongestSection", "m", "Longest bonding section length", "", "For cross-bonding calculations")
    add(T, "LongerSection", "m", "Longer bonding section length", "", "For cross-bonding calculations")
    add(T, "ShortestSection", "m", "Shortest bonding section length", "", "For cross-bonding calculations")

    # Thermal capacitance (for transient rating per IEC 60853)
    add(T, "BETAC", "J/(m³·K)", "Conductor volumetric thermal capacitance", "", "IEC 60853 transient parameter")
    add(T, "SHC", "J/(kg·K)", "Conductor specific heat capacity", "", "IEC 60853 transient parameter")
    add(T, "BETAS", "J/(m³·K)", "Sheath volumetric thermal capacitance", "", "IEC 60853 transient parameter")
    add(T, "SHS", "J/(kg·K)", "Sheath specific heat capacity", "", "IEC 60853 transient parameter")
    add(T, "BETAT", "J/(m³·K)", "Reinforcing tape volumetric thermal capacitance", "", "IEC 60853 transient parameter")
    add(T, "SHT", "J/(kg·K)", "Reinforcing tape specific heat capacity", "", "IEC 60853 transient parameter")
    add(T, "BETAK", "J/(m³·K)", "Concentric neutral volumetric thermal capacitance", "", "IEC 60853 transient parameter")
    add(T, "SHK", "J/(kg·K)", "Concentric neutral specific heat capacity", "", "IEC 60853 transient parameter")
    add(T, "BETAA", "J/(m³·K)", "Armour volumetric thermal capacitance", "", "IEC 60853 transient parameter")
    add(T, "SHA", "J/(kg·K)", "Armour specific heat capacity", "", "IEC 60853 transient parameter")

    # Misc flags and rarely-used fields
    add(T, "Dimension_correction_flag", "", "Dimension correction flag", "0,2", "Purpose unclear")
    add(T, "Duct_custom_U", "", "Custom duct parameter U (unused)", "", "Not populated")
    add(T, "Duct_custom_V", "", "Custom duct parameter V (unused)", "", "Not populated")
    add(T, "Duct_custom_Y", "", "Custom duct parameter Y (unused)", "", "Not populated")
    add(T, "JACKETMAT_SLTYPE", "", "Per-core jacket material code (3-core)", JACKMAT_ENUM, "Jacket around each individual core")
    add(T, "JACKETCNAM_SLTYPE", "", "Per-core custom jacket material name", "", "Text label")
    add(T, "RHJ_SLTYPE", "K.m/W", "Per-core custom jacket thermal resistivity", "", "For individual core jackets")
    add(T, "JACKET_THI_SLTYPE", "m", "Per-core jacket thickness", "", "Only for specific cable types (e.g. city cable)")
    add(T, "GROUP_CABLE_NAME", "", "Cable group/folder name", "", "Organizational folder in CYMCAP UI")
    add(T, "K_SpecificInsulationResis", "M.ohm.km", "K factor for insulation resistance", "", "Specific insulation resistance")
    add(T, "Frequency_flag", "", "Frequency calculation flag (unused)", "", "Not populated")
    add(T, "Frequency", "Hz", "System frequency (cable-level override)", "", "Not populated; typically set at study level")
    add(T, "INSSYS", "", "Insulation system code", "1,2", "Purpose unclear")
    add(T, "MILWIRECON", "", "Milliken wire construction type", "1=Insulated, 2=Unidirectional, 3=Bidirectional", "For Milliken conductors")
    add(T, "MEDIUM_DUCT_FLAG", "", "Medium/duct flag (unused)", "", "Not populated")
    add(T, "MAGNETIC_PROPERTY_FLAG", "", "Magnetic property flag (unused)", "", "Not populated")
    add(T, "OUTER_CABLE_PHASE_FLAG", "", "Outer cable phase flag (unused)", "", "Not populated")
    add(T, "CONDSHLDMAT", "", "Conductor shield material code", CONDSHLDMAT_ENUM, "Maps to ConductorShieldMaterial enum")
    add(T, "RHCS", "K.m/W", "Custom conductor shield thermal resistivity", "", "Used when CONDSHLDMAT=-1 (custom)")
    add(T, "CONDSHLDCNAM", "", "Custom conductor shield material name", "", "Text label")
    add(T, "CORE_LENGTH_OF_LAY", "m", "Core length of lay (3-core cables)", "", "Helical lay length of individual cores")
    add(T, "LENGTH_OF_PITCH", "m", "Corrugated sheath pitch length", "", "Corrugation pitch for corrugated sheaths")
    add(T, "ARMROUNDWI", "", "Armour wire shape flag", "0=N/A, 1=Round Wire, 2=Flat Straps", "Wire vs strap armour")
    add(T, "ARMSTRAPLG", "m", "Armour strap width", "", "Width of flat strap armour")
    add(T, "ACDC_FLAG", "", "AC/DC flag", ACDC_ENUM, "Maps to AcDcFlag enum")

    # Second armour layer fields
    for suffix, label in [("ARMREIN2", "Second armour type code"), ("ARMBED2", "Second armour bedding material"),
                          ("ARMBEDTH2", "Second armour bedding thickness"), ("ARMBED_DIAM2", "Second armour bedding diameter"),
                          ("ARMWIRDI2", "Second armour wire diameter"), ("WIARMDIA2", "Second armour outer diameter"),
                          ("TAPARMAR2", "Second armour tape area"), ("ARMWIRLE2", "Second armour wire lay length"),
                          ("TAPARMDI2", "Second armour tape diameter"), ("ARMWIRNU2", "Second armour wire count"),
                          ("ARMROUNDWI2", "Second armour wire shape"), ("ARMSTRAPLG2", "Second armour strap width"),
                          ("ArmourEmbeddedPEInnerTh2", "Second armour PE inner thickness"),
                          ("ArmourEmbeddedAluminumTh2", "Second armour aluminium thickness"),
                          ("ArmourEmbeddedPEOuterTh2", "Second armour PE outer thickness"),
                          ("PERMEAB2", "Second armour permeability"), ("RHA2", "Second armour resistivity"),
                          ("ALFAA2", "Second armour temp coefficient"), ("BETAA2", "Second armour thermal capacitance"),
                          ("SHA2", "Second armour specific heat"), ("RHB2", "Second armour bedding thermal resistivity"),
                          ("AME2", "Second armour AME parameter"), ("AMT2", "Second armour AMT parameter"),
                          ("GAMMA2", "Second armour GAMMA parameter"),
                          ("ARMBEDCNAM2", "Second armour bedding custom material name"),
                          ("ARMOURCNAM2", "Second armour custom material name")]:
        add(T, suffix, "", f"{label} (second armour layer — not populated)", "", "For double-armoured cables; not used in sample DB")

    add(T, "CITYCABLE_FLAG", "", "Retrofitted city cable flag", "1=City cable (e.g. CIGRE Case 9)", "Indicates special city cable configuration")
    add(T, "ShCctSecondArmourData", "", "Short-circuit second armour data (CSV string)", "", "For double-armoured cables")

    # ===================================================================
    # CABLE_LAYERS (11 columns) — fully documented
    # ===================================================================
    T = "CABLE_LAYERS"
    add(T, "IDCAB", "", "Cable ID — foreign key to CABLES", "", "Links layer to parent cable")
    add(T, "LAYERID", "", "Layer index (0-based from conductor outward)", "", "Ordering of layers radially outward")
    add(T, "LAYERTYPE", "", "Layer type category", "e.g. Conductor, Insulation, Shield, Sheath, Jacket, Armour, Bedding, Serving", "Identifies which cable component")
    add(T, "LAYERNAME", "", "Layer display name", "", "Human-readable layer name from CYMCAP")
    add(T, "MATERIAL", "", "Material identifier or code", "", "Material code or name string")
    add(T, "CUSTOM_MATERIAL_NAME", "", "Custom material name (when material is 'Custom')", "", "User-defined material label")
    add(T, "THICKNESS", "m", "Layer thickness", "", "Radial thickness in metres")
    add(T, "DIAMETER", "m", "Layer outer diameter", "", "Outer diameter after this layer")
    add(T, "THERMAL_RESIS", "K.m/W", "Thermal resistivity of layer material", "", "For thermal circuit calculations")
    add(T, "INTERNAL_THREE_CORE", "", "Internal to 3-core assembly flag", "0=No, 1=Yes", "Whether layer is inside the 3-core bundle")
    add(T, "TAG", "", "Record tag", "", "Internal flag")

    # ===================================================================
    # CABLES_INST (74 columns) — 14 core + 60 GUI/label
    # ===================================================================
    T = "CABLES_INST"
    add(T, "IdStudy", "", "Study ID — foreign key to STUDYHED", "", "Links installation to parent study")
    add(T, "Execuno", "", "Execution number within study", "", "Sub-case identifier")
    add(T, "CabIndex", "", "Cable position index (0-based)", "", "Order of cable within installation layout")
    add(T, "FeederId", "", "Feeder identifier", "", "Circuit/feeder grouping label")
    add(T, "Circuit", "", "Circuit number", "", "1-based circuit index")
    add(T, "PhaseId", "", "Phase identifier", "", "e.g. 'A', 'B', 'C' or '1', '2', '3'")
    add(T, "IdCab", "", "Cable ID — foreign key to CABLES", "", "Which cable design is installed here")
    add(T, "XL", "m", "Horizontal position", "", "X coordinate of cable centre")
    add(T, "YL", "m", "Vertical position (depth, negative = underground)", "", "Y coordinate (depth below surface)")
    add(T, "AmpTemp", "°C", "Maximum conductor temperature for ampacity", "", "Temperature limit for this installation point")
    add(T, "RefFrac", "", "Reference fraction", "", "Fraction of reference current")
    add(T, "LoadFactor", "", "Load factor", "", "Ratio of average to peak load")
    add(T, "Layout", "", "Layout identifier", "", "Layout configuration string")
    add(T, "DuctNo", "", "Duct number (0 = direct buried)", "", "Which duct in a duct bank (0 if none)")

    # Remaining CABLES_INST columns (GUI label/display and extended fields)
    add(T, "Positions", "", "Cable positions data (serialized)", "", "Internal layout data")
    add(T, "LayerIndexNo", "", "Soil/backfill layer index", "", "Which thermal layer this cable is in")
    add(T, "XL_CenterOfDuctsInTrefoil", "m", "X centre of trefoil duct group", "", "For trefoil duct arrangements")
    add(T, "YL_CenterOfDuctsInTrefoil", "m", "Y centre of trefoil duct group", "", "For trefoil duct arrangements")
    add(T, "Label_X", "px", "Label X position (GUI)", "", "GUI display coordinate")
    add(T, "Label_Y", "px", "Label Y position (GUI)", "", "GUI display coordinate")
    add(T, "Label_Width", "px", "Label width (GUI)", "", "GUI display size")
    add(T, "Label_Height", "px", "Label height (GUI)", "", "GUI display size")
    add(T, "Label_AMP_A", "", "Label ampacity phase A text (GUI)", "", "GUI display")
    add(T, "Label_TEMP_A", "", "Label temperature phase A text (GUI)", "", "GUI display")
    add(T, "Label_AMP_B", "", "Label ampacity phase B text (GUI)", "", "GUI display")
    add(T, "Label_TEMP_B", "", "Label temperature phase B text (GUI)", "", "GUI display")
    add(T, "Label_AMP_C", "", "Label ampacity phase C text (GUI)", "", "GUI display")
    add(T, "Label_TEMP_C", "", "Label temperature phase C text (GUI)", "", "GUI display")
    add(T, "Label_Align_Hor", "", "Label horizontal alignment (GUI)", "", "GUI display")
    add(T, "Label_Align_Ver", "", "Label vertical alignment (GUI)", "", "GUI display")
    add(T, "Label_Bgd_Color", "", "Label background colour (GUI)", "", "GUI display")
    add(T, "Label_Font_Color", "", "Label font colour (GUI)", "", "GUI display")
    add(T, "Label_Connected_Line", "", "Label connector line flag (GUI)", "", "GUI display")
    add(T, "NbCables_per_duct", "", "Number of cables per duct", "", "For multiple cables in single duct")
    add(T, "YEARLY_LOADFACTOR", "", "Yearly load factor", "", "Annual average/peak ratio")
    add(T, "WEEKLY_LOADFACTOR", "", "Weekly load factor", "", "Weekly average/peak ratio")
    add(T, "ANGLE", "°", "Cable installation angle", "", "Rotation angle of cable group")
    add(T, "SUBLAYOUT", "", "Sub-layout identifier", "", "Layout subdivision")
    add(T, "LABEL_CUSTOMIZE_VERSION", "", "Label customization version (GUI)", "", "GUI version tracking")
    add(T, "LABEL_TEXT_LANGUAGE", "", "Label text language (GUI)", "", "GUI localisation")
    add(T, "LABEL_SHOW_CIRCUITNO", "", "Show circuit number in label (GUI)", "", "GUI display toggle")
    add(T, "LABEL_CIRCUITNO_ORDER", "", "Circuit number display order (GUI)", "", "GUI display order")
    add(T, "LABEL_SHOW_FEEDERID", "", "Show feeder ID in label (GUI)", "", "GUI display toggle")
    add(T, "LABEL_FEEDERID_ORDER", "", "Feeder ID display order (GUI)", "", "GUI display order")
    add(T, "LABEL_SHOW_CABLEID", "", "Show cable ID in label (GUI)", "", "GUI display toggle")
    add(T, "LABEL_CABLEID_ORDER", "", "Cable ID display order (GUI)", "", "GUI display order")
    add(T, "LABEL_SHOW_PHASE", "", "Show phase in label (GUI)", "", "GUI display toggle")
    add(T, "LABEL_PHASE_ORDER", "", "Phase display order (GUI)", "", "GUI display order")
    add(T, "LABEL_SHOW_LOAD_FACTOR", "", "Show load factor in label (GUI)", "", "GUI display toggle")
    add(T, "LABEL_LOAD_FACTOR_ORDER", "", "Load factor display order (GUI)", "", "GUI display order")
    add(T, "LABEL_SHOW_ANGLE", "", "Show angle in label (GUI)", "", "GUI display toggle")
    add(T, "LABEL_ANGLE_ORDER", "", "Angle display order (GUI)", "", "GUI display order")
    add(T, "LABEL_SHOW_FREQUENCY", "", "Show frequency in label (GUI)", "", "GUI display toggle")
    add(T, "LABEL_FREQUENCY_ORDER", "", "Frequency display order (GUI)", "", "GUI display order")
    add(T, "LABEL_FONT_SIZE", "pt", "Label font size (GUI)", "", "GUI display")
    add(T, "LABEL_SHOW_CABLE_VOLTAGE", "", "Show cable voltage in label (GUI)", "", "GUI display toggle")
    add(T, "LABEL_CABLE_VOLTAGE_ORDER", "", "Cable voltage display order (GUI)", "", "GUI display order")
    add(T, "LABEL_SHOW_CABLE_SIZE", "", "Show cable size in label (GUI)", "", "GUI display toggle")
    add(T, "LABEL_CABLE_SIZE_ORDER", "", "Cable size display order (GUI)", "", "GUI display order")
    add(T, "AMPTEMP_EXTRA", "", "Extra ampacity/temperature data", "", "Additional calc data")
    add(T, "RTTR_STATUSFLAG", "", "Real-time thermal rating status flag", "", "RTTR module status")
    add(T, "TREFOIL_DISTANCE", "m", "Distance between trefoil cables", "", "Spacing for trefoil arrangement")
    add(T, "MCPP_STATUS_FLAG", "", "MCPP module status flag", "", "Multi-cable per phase status")
    add(T, "REFER_CABLE_DEPTH_FLAG", "", "Reference cable depth flag", "", "Depth reference method")
    add(T, "MAX_INSU_TEMP_DROP", "°C", "Maximum insulation temperature drop", "", "Insulation thermal constraint")
    add(T, "XL_3D", "m", "3D X position", "", "For 3D visualisation")
    add(T, "ZL_3D", "m", "3D Z position", "", "For 3D visualisation")
    add(T, "SHOWAXIS_3D_FLAG", "", "Show 3D axis flag (GUI)", "", "GUI display toggle")
    add(T, "MULTIPTS_3D_X", "", "3D multi-point X coordinates", "", "For 3D cable routing")
    add(T, "MULTIPTS_3D_Y", "", "3D multi-point Y coordinates", "", "For 3D cable routing")
    add(T, "MULTIPTS_3D_Z", "", "3D multi-point Z coordinates", "", "For 3D cable routing")
    add(T, "MULTIPTS_3D_BRADIUS", "m", "3D multi-point bend radius", "", "For 3D cable routing")
    add(T, "ORIGIN_STARTING_PTS_XYZ", "", "3D origin starting points", "", "For 3D cable routing")
    add(T, "RELATIVE_POS_3D", "", "3D relative position", "", "For 3D cable routing")

    # ===================================================================
    # STUDYHED (184 columns) — comprehensive
    # ===================================================================
    T = "STUDYHED"
    add(T, "IDSTUDY", "", "Study ID — unique identifier", "", "Primary key, VARCHAR")
    add(T, "EXECUNO", "", "Execution number within study", "", "Sub-case identifier within study")
    add(T, "STUDYTITLE", "", "Study title", "", "User-entered study description")
    add(T, "EXECDATE", "", "Execution date", "", "Date of calculation run")
    add(T, "COMMENTS", "", "Study comments", "", "User-entered notes")
    add(T, "EXECUTITLE", "", "Execution title", "", "User-entered sub-case description")
    add(T, "TAG", "", "Record tag", "", "Internal flag")
    add(T, "TAG2", "", "Secondary record tag", "", "Internal flag")
    add(T, "MODECALC", "", "Calculation mode", MODECALC_ENUM, "Which calculation method to use")
    add(T, "NCABLE", "", "Number of cables in study", "", "Count of cable positions")
    add(T, "NTYPE", "", "Number of cable types", "", "Distinct cable designs used")
    add(T, "AMBIENT", "°C", "Ambient soil/air temperature", "", "IEC 60287 ambient temperature")
    add(T, "RHOS", "K.m/W", "Native soil thermal resistivity", "", "IEC 60287 soil thermal resistivity")
    add(T, "IREFCA", "A", "Reference cable current", "", "Initial/reference current for iteration")
    add(T, "XB", "m", "Drawing X boundary (GUI)", "", "Cross-section view boundary")
    add(T, "YB", "m", "Drawing Y boundary (GUI)", "", "Cross-section view boundary")
    add(T, "ALB", "", "Drawing scale parameter (GUI)", "", "Cross-section view scale")
    add(T, "AXB", "", "Drawing axis parameter (GUI)", "", "Cross-section view axis")
    add(T, "RHOB", "K.m/W", "Backfill thermal resistivity", "", "Thermal resistivity of installed backfill")
    add(T, "CABINAIR", "", "Cable in air flag", "", "Whether any cables are in air")
    add(T, "EXTHEAT", "W/m", "External heat source rate", "", "External heat dissipation per unit length")
    add(T, "TEMTOL", "°C", "Temperature convergence tolerance", "", "Solver convergence criterion")
    add(T, "AMPTOL", "A", "Ampacity convergence tolerance", "", "Solver convergence criterion")
    add(T, "MAXNIT", "", "Maximum iterations", "", "Solver iteration limit")
    add(T, "TRTOL", "", "Transient tolerance", "", "Transient solver convergence criterion")
    add(T, "TRSTEP", "s", "Transient time step", "", "Transient solver time step")
    add(T, "TRMAXNIT", "", "Transient max iterations", "", "Transient solver iteration limit")
    add(T, "FRACT", "", "Load fraction", "", "Fraction of full load for calculation")
    add(T, "FREQ", "Hz", "System frequency", "", "50 or 60 Hz for AC loss calculations")
    add(T, "PRINTI", "", "Print intermediate results flag", "", "Output control")
    add(T, "PRINT", "", "Print results flag", "", "Output control")
    add(T, "EXRESMO", "", "Extended results mode", "", "Output detail level")
    add(T, "CAAIRAR", "", "Cable in air arrangement", "", "Configuration for cables in air")
    add(T, "SHADED", "", "Shaded/direct sunlight flag", "", "Solar radiation exposure")
    add(T, "SUNINTEN", "W/m²", "Solar radiation intensity", "", "IEC 60287 solar radiation")
    add(T, "SUNABSO", "", "Solar absorptivity coefficient", "", "Cable surface solar absorptivity")
    add(T, "HEORTE", "", "Heat source or temperature boundary", "", "Boundary condition type")
    add(T, "HEATEMP", "°C", "Heat source temperature", "", "Fixed temperature boundary")
    add(T, "XLHEAT", "m", "Heat source X position", "", "X coordinate of external heat source")
    add(T, "YLHEAT", "m", "Heat source Y position", "", "Y coordinate of external heat source")
    add(T, "TR_HEATSRC", "", "Transient heat source flag", "", "Whether heat source varies with time")
    add(T, "HEATSRC_ID", "", "Heat source library ID", "", "FK to HEATSRCLIB")
    add(T, "HEATSRC_SF", "", "Heat source scale factor", "", "Multiplier on library heat rate")
    add(T, "DIAHEAT", "m", "Heat source diameter", "", "Physical size of external heat source")
    add(T, "AMBAIR", "°C", "Ambient air temperature", "", "For cable sections in air")
    add(T, "CONVCO", "W/(m²·K)", "Convection coefficient", "", "For cable-in-air heat transfer")
    add(T, "DUCT", "", "Duct flag", "", "Whether installation uses ducts")
    add(T, "MIGFLAG", "", "Moisture migration flag", "", "Whether to model dry-out zone")
    add(T, "RHDX", "K.m/W", "Dry zone thermal resistivity", "", "Soil resistivity in dried-out zone")
    add(T, "ISOT", "", "Isothermal surface flag", "", "Boundary condition setting")
    add(T, "RBOX", "", "Result box display flag (GUI)", "", "GUI display option")
    add(T, "TRANOPT", "", "Transient option", "", "Transient calculation sub-option")
    add(T, "IOP", "", "Iterative optimisation parameter", "", "Solver tuning")
    add(T, "STARTVAL", "", "Parameter sweep start value", "", "For parametric studies")
    add(T, "ENDVAL", "", "Parameter sweep end value", "", "For parametric studies")
    add(T, "STEP", "", "Parameter sweep step", "", "For parametric studies")
    add(T, "ENERGRATE", "", "Energy rate (economic analysis)", "", "Cost per kWh")
    add(T, "DISCRATE", "%", "Discount rate (economic analysis)", "", "NPV discount rate")
    add(T, "ENERCOST", "", "Energy cost (economic analysis)", "", "Total energy cost")
    add(T, "DEMDCOST", "", "Demand cost (economic analysis)", "", "Peak demand charge")
    add(T, "LINCOST", "", "Linear cost (economic analysis)", "", "Per-metre cable cost")
    add(T, "ECONLIFE", "years", "Economic life (economic analysis)", "", "Design lifetime")
    add(T, "INITINVEST", "", "Initial investment (economic analysis)", "", "Capital cost")
    add(T, "LENGTH", "m", "Cable route length", "", "Total cable route length")
    add(T, "CHARGING", "A", "Charging current", "", "Capacitive charging current")
    add(T, "INITCURNT", "A", "Initial current", "", "Starting current for time-varying analysis")
    add(T, "LOADGROWTH", "%/yr", "Load growth rate", "", "Annual load growth")
    add(T, "ULTIMLF", "", "Ultimate load factor", "", "Long-term target load factor")
    add(T, "GROWTHYEAR", "years", "Growth period", "", "Years over which load grows")
    add(T, "WIND", "m/s", "Wind speed", "", "For cable-in-air convection")
    add(T, "WLENGTH", "m", "Wavelength parameter", "", "For electromagnetic calculation")
    add(T, "PERMTR", "", "Perimeter parameter", "", "For tunnel/trough calculations")
    add(T, "INSTTYPE", "", "Installation type code", INSTTYPE_ENUM, "Underground, submarine, air, etc.")
    add(T, "DUCTBANK", "", "Duct bank flag", "", "Whether duct bank is used")
    add(T, "DUCTBANKID", "", "Duct bank library ID reference", "", "Links to DUCTLIB table")
    add(T, "DUCTYPOS", "m", "Duct bank Y position (depth)", "", "Depth to duct bank reference point")
    add(T, "RHD", "K.m/W", "Duct material thermal resistivity", "", "Study-level duct thermal resistivity")
    add(T, "DUCTNAME", "", "Duct bank name", "", "Display name")
    add(T, "OFFSETX", "m", "Duct bank X offset", "", "Horizontal offset of duct bank")
    add(T, "OFFSETY", "m", "Duct bank Y offset", "", "Vertical offset of duct bank")
    add(T, "OFFSETX2", "m", "Duct bank X offset (secondary)", "", "For two-bank configurations")
    add(T, "OFFSETY2", "m", "Duct bank Y offset (secondary)", "", "For two-bank configurations")
    add(T, "DISTX", "m", "Duct X spacing", "", "Horizontal distance between ducts")
    add(T, "DISTY", "m", "Duct Y spacing", "", "Vertical distance between ducts")
    add(T, "TOTROW", "", "Total duct rows", "", "Duct bank row count")
    add(T, "TOTCOL", "", "Total duct columns", "", "Duct bank column count")
    add(T, "DUPIINDI", "m", "Duct inside diameter (study-level)", "", "Study override for duct ID")
    add(T, "DUPIOTDI", "m", "Duct outside diameter (study-level)", "", "Study override for duct OD")
    add(T, "MOISTURE", "%", "Critical moisture content", "", "Threshold for moisture migration")
    add(T, "TRANSIENT", "", "Transient analysis flag", "", "Whether transient is enabled")
    add(T, "OPTICALC", "", "Optimisation calculation flag", "", "Whether to optimise")
    add(T, "OPTICIRCNO", "", "Optimisation circuit number", "", "Which circuit to optimise")
    add(T, "CHOTE_COMPATIBLE", "", "CHOTE compatibility flag", "", "Compatibility with CHOTE format")
    add(T, "LOCKEDBY", "", "Locked by user", "", "Multi-user locking field")
    add(T, "LOCKTSTAMP", "", "Lock timestamp", "", "Multi-user locking timestamp")
    add(T, "AutoRefCircuitFlag", "", "Auto-reference circuit flag", "", "Automatic reference circuit selection")
    add(T, "CHOTE_TIMELINE_DATE", "", "CHOTE timeline date", "", "Timeline reference date")
    add(T, "HEATSOURCE_INSULATION_FLAG", "", "Heat source insulation flag", "", "Whether heat source has insulation")
    add(T, "INSUDIAM_HEATSOURCE", "m", "Heat source insulation diameter", "", "Insulation OD of heat source")
    add(T, "THRESIS_HEATSOURCE", "K.m/W", "Heat source insulation thermal resistivity", "", "Insulation of heat source")
    add(T, "POWERFACTOR", "", "Power factor", "", "For AC power calculation")
    add(T, "LAGGING_LEADING_FLAG", "", "Lagging/leading power factor flag", "", "0=Lagging, 1=Leading")
    add(T, "TUNNEL_IMAGE_CHOSEN", "", "Tunnel image selection (GUI)", "", "GUI display")
    add(T, "TUNNEL_CABLETYPE_CHOSEN", "", "Tunnel cable type selection", "", "Cable arrangement in tunnel")
    add(T, "TUNNEL_DIAM", "m", "Tunnel diameter", "", "For tunnel installation")
    add(T, "TUNNEL_ESTHECTIC_PARAM", "", "Tunnel aesthetic parameter (GUI)", "", "GUI display")
    add(T, "CYCLIC_LOADING", "", "Cyclic loading flag", "", "Whether cyclic loading is applied")
    add(T, "EMF_XMIN", "m", "EMF calculation X minimum", "", "EMF field calculation range")
    add(T, "EMF_XMAX", "m", "EMF calculation X maximum", "", "EMF field calculation range")
    add(T, "EMF_XSTEP", "m", "EMF calculation X step", "", "EMF field calculation resolution")
    add(T, "EMF_YVALUES", "", "EMF Y values (serialized)", "", "EMF field Y coordinates")
    add(T, "TF_LEVEL_VALUES", "", "Temperature field contour levels", "", "Thermal field visualisation")
    add(T, "TF_LEVEL_COLORS", "", "Temperature field contour colours", "", "Thermal field visualisation")
    add(T, "TF_LABEL_SPACING", "", "Temperature field label spacing", "", "Thermal field visualisation")
    add(T, "TF_RESOLUTION", "", "Temperature field resolution", "", "Thermal field grid resolution")
    add(T, "TF_ZOOMOUT_POSITION", "", "Temperature field zoom position (GUI)", "", "GUI display")
    add(T, "TG_FRAME_LEFTPOS", "m", "Trough/gutter frame left position", "", "Trough geometry")
    add(T, "TG_FRAME_TOPPOS", "m", "Trough/gutter frame top position", "", "Trough geometry")
    add(T, "TG_FRAME_WIDTH", "m", "Trough/gutter frame width", "", "Trough geometry")
    add(T, "TG_FRAME_HEIGHT", "m", "Trough/gutter frame height", "", "Trough geometry")
    add(T, "TG_FRAME_THRESIS", "K.m/W", "Trough/gutter frame thermal resistivity", "", "Trough material")
    add(T, "TG_INSIDE_TEXTURE", "", "Trough inside texture (GUI)", "", "GUI display")
    add(T, "TG_FRAME_TEXTURE", "", "Trough frame texture (GUI)", "", "GUI display")
    add(T, "CROSSINGS_FLAG", "", "Cable crossings flag", "", "Whether cable crossings are modelled")
    add(T, "HEATSOURCE_COLD_OR_HOT", "", "Heat source hot/cold flag", "", "Whether source heats or cools")
    add(T, "CASINGID", "", "Casing library ID", "", "FK to CASINGLIB")
    add(T, "MULTCASING_INSTALLATION_TYPE", "", "Multiple casing installation type", "", "Configuration")
    add(T, "DEPTH_OF_SEABED", "m", "Depth of seabed", "", "For submarine cable installation")
    add(T, "EARTH_RESISTIVITY", "ohm.m", "Earth electrical resistivity", "", "For grounding calculations")
    add(T, "COMMON_FREQUENCY_FLAG", "", "Common frequency flag", "", "Whether all cables share same frequency")
    add(T, "CABLE_RESISTANCE_FLAG", "", "Cable resistance calculation flag", "", "How AC resistance is computed")
    add(T, "USE_ENLARGE_BOUNDARIES", "", "Enlarged boundaries flag", "", "FEM boundary extension")
    add(T, "ENLARGE_BOUNDARIES_VALUE", "", "Enlarged boundaries value", "", "FEM boundary multiplier")
    add(T, "ELECTRICAL_INTERATIVE_BETWEEN_CCTS", "", "Electrical interaction between circuits flag", "", "Inter-circuit coupling")
    add(T, "TG_SHALLOW_TROUGHS_METHOD", "", "Shallow troughs calculation method", "", "IEC method selection")
    add(T, "CUSTOM_GENERAL_RATINGS_FLAG", "", "Custom general ratings flag", "", "Override standard ratings")
    add(T, "CABLEPARTS_TO_ITERATE_FLAG", "", "Cable parts iteration flag", "", "Which parts to iterate on")
    add(T, "DUCTBANK_CUSTOM_U", "", "Custom duct bank parameter U", "", "Custom thermal coefficient")
    add(T, "DUCTBANK_CUSTOM_V", "", "Custom duct bank parameter V", "", "Custom thermal coefficient")
    add(T, "DUCTBANK_CUSTOM_Y", "", "Custom duct bank parameter Y", "", "Custom thermal coefficient")
    add(T, "STD_NOMINAL_DUCT", "", "Standard nominal duct size", "", "Nominal duct designation")
    add(T, "CLEARANCES_MODE", "", "Clearances calculation mode", "", "How clearances are computed")
    add(T, "TF_STEP_RESOLUTION", "", "Temperature field step resolution", "", "FEM mesh refinement")
    add(T, "MEDIUM_DUCT_FLAG", "", "Medium duct flag", "", "Duct fill medium type")
    add(T, "MAGNETIC_PROPERTY_FLAG", "", "Magnetic property calculation flag", "", "Include magnetic effects")
    add(T, "COND_TEMP_OVERLIMIT_FLAG", "", "Conductor temperature over-limit flag", "", "Warning/constraint")
    add(T, "TUNNEL_VENTILATED_FLAG", "", "Tunnel ventilation flag", "", "Whether tunnel has forced air")
    add(T, "AIR_VELOCITY_IN_TUNNEL", "m/s", "Air velocity in tunnel", "", "Forced ventilation speed")
    add(T, "TEMP_OF_AIR_IN_TUNNEL", "°C", "Air temperature in tunnel", "", "Inlet air temperature")
    add(T, "LENGTH_OF_TUNNEL", "m", "Tunnel length", "", "Total tunnel length")
    add(T, "SHAPE_OF_TUNNEL", "", "Tunnel cross-section shape", "", "Circular, rectangular, etc.")
    add(T, "INNER_DIAM_OF_TUNNEL", "m", "Inner diameter of tunnel", "", "For circular tunnels")
    add(T, "IEC_SCOPE_EXTENDED_FLAG", "", "IEC extended scope flag", "", "Whether extended IEC methods are used")
    add(T, "RAC_FACTOR_MAGNETIC_DUCT", "", "AC resistance factor for magnetic duct", "", "Correction factor")
    add(T, "DUCT_FILLING_TH_RESIS", "K.m/W", "Duct filling thermal resistivity", "", "Thermal resistivity of duct fill")
    add(T, "MDB_DUCTS_TREFOIL_RECALC_FLAG", "", "Duct trefoil recalculation flag", "", "Force recalculation")
    add(T, "COMPUTE_SOLAR_RADIATION_FLAG", "", "Compute solar radiation flag", "", "Whether to include solar heating")
    add(T, "SUNINTEN_PERPENDICULAR", "W/m²", "Perpendicular solar radiation intensity", "", "Solar flux normal to cable")
    add(T, "ANGLE_INCIDENCE", "°", "Angle of solar incidence", "", "Sun angle to cable surface")
    add(T, "FACTOR_RAC_DUE_TO_MAGNETIC_ARMOUR_FLAG", "", "RAC magnetic armour factor flag", "", "IEC correction flag")
    add(T, "FACTOR_RAC_DUE_TO_MAGNETIC_DUCT_FLAG", "", "RAC magnetic duct factor flag", "", "IEC correction flag")
    add(T, "FACTOR_ON_LAMBDA1_DUE_TO_MAGNETIC_ARMOUR_FLAG", "", "Lambda1 magnetic armour factor flag", "", "IEC correction flag")
    add(T, "FACTOR_ON_LAMBDA1_DUE_TO_MAGNETIC_DUCT_FLAG", "", "Lambda1 magnetic duct factor flag", "", "IEC correction flag")
    add(T, "FACTOR_ON_LAMBDA2_ARMOUR_TO_LAMBDA1_CIRCULATING_FLAG", "", "Lambda2-to-lambda1 circulating factor flag", "", "IEC correction flag")
    add(T, "LAMBDA1_EDDY_COMPUTATION_FLAG", "", "Lambda1 eddy current computation flag", "", "IEC loss calculation method")
    add(T, "LAMBDA2_PIPE_COMPUTATION_FLAG", "", "Lambda2 pipe computation flag", "", "IEC loss calculation method")
    add(T, "WD_COMPUTATION_FLAG", "", "Wd (dielectric loss) computation flag", "", "IEC loss calculation method")
    add(T, "T4_COMPUTATION_FLAG", "", "T4 (external thermal resistance) computation flag", "", "IEC method")
    add(T, "CUSTOM_FACTOR_RAC_DUE_TO_MAGNETIC_ARMOUR", "", "Custom RAC factor for magnetic armour", "", "User override value")
    add(T, "CUSTOM_FACTOR_RAC_DUE_TO_MAGNETIC_DUCT", "", "Custom RAC factor for magnetic duct", "", "User override value")
    add(T, "CUSTOM_FACTOR_ON_LAMBDA1_DUE_TO_MAGNETIC_ARMOUR", "", "Custom lambda1 factor for magnetic armour", "", "User override value")
    add(T, "CUSTOM_FACTOR_ON_LAMBDA1_DUE_TO_MAGNETIC_DUCT", "", "Custom lambda1 factor for magnetic duct", "", "User override value")
    add(T, "CUSTOM_FACTOR_ON_LAMBDA2_ARMOUR_TO_LAMBDA1_CIRCULATING", "", "Custom lambda2-to-lambda1 circulating factor", "", "User override value")
    add(T, "GENERAL_GUIDANCE_FLAG", "", "General guidance/notes flag", "", "Whether guidance notes are shown")
    add(T, "T2_COMPUTATION_FLAG", "", "T2 (between sheath and armour) computation flag", "", "IEC method")
    add(T, "ENVIRONMENT_CONFIG_ID", "", "Environment configuration ID", "", "FK to GLOBAL_ENVIRONMENT_PARAMETERS")
    add(T, "ENVIRONMENT_OFFSET_AMBIENT", "°C", "Ambient temperature offset", "", "Adjustment to environment config ambient")
    add(T, "CABIN_JTUBE_FLAG", "", "Cable in J-tube flag", "", "For J-tube installations")
    add(T, "JTUBE_LENGTH_ABOVE_WATER", "m", "J-tube length above water", "", "Exposed J-tube section")
    add(T, "CAB_IMMERSED_IN_WATER_FLAG", "", "Cable immersed in water flag", "", "Submarine section flag")
    add(T, "HIGH_VOLTAGE_DC_CABLE_FLAG", "", "HVDC cable flag", "", "For DC-specific calculations")
    add(T, "TUNNELWALL_ADVANCED_FLAG", "", "Advanced tunnel wall modelling flag", "", "Detailed wall thermal model")
    add(T, "TUNNELWALL_THICKNESS", "m", "Tunnel wall thickness", "", "For wall thermal resistance")
    add(T, "TUNNELWALL_THERMALRESIS", "K.m/W", "Tunnel wall thermal resistivity", "", "Wall material property")
    add(T, "NUMBERLAYERS", "", "Number of soil/backfill layers", "", "For multi-layer thermal environment")

    # ===================================================================
    # GLOBAL_ENVIRONMENT_PARAMETERS (11 columns) — fully documented
    # ===================================================================
    T = "GLOBAL_ENVIRONMENT_PARAMETERS"
    add(T, "ENVIRONMENT_CONFIG_ID", "", "Configuration ID — unique identifier", "", "Primary key, VARCHAR")
    add(T, "ENVIRONMENT_CONFIG_NAME", "", "Configuration name", "", "User-entered label")
    add(T, "AMBIENT", "°C", "Ambient soil temperature", "", "Default ambient for studies using this config")
    add(T, "RHOS", "K.m/W", "Native soil thermal resistivity", "", "Default soil thermal resistivity")
    add(T, "DEPTH_FLAG", "", "Depth calculation flag", "", "How burial depth is handled")
    add(T, "ISOT", "", "Isothermal surface flag", "", "Boundary condition setting")
    add(T, "AMBAIR", "°C", "Ambient air temperature", "", "For cable sections in air")
    add(T, "MIGFLAG", "", "Moisture migration flag", "", "Whether to model dry-out zone")
    add(T, "RHDX", "K.m/W", "Dry zone thermal resistivity", "", "Soil resistivity in dried-out zone")
    add(T, "MOISTURE", "%", "Critical moisture content", "", "Threshold for moisture migration")
    add(T, "TAG", "", "Record tag", "", "Internal flag")

    # ===================================================================
    # LOADWIN (10 columns) — fully documented
    # ===================================================================
    T = "LOADWIN"
    add(T, "KEYLOAD", "", "Load window ID — unique identifier", "", "Primary key, VARCHAR")
    add(T, "TITLE", "", "Load window title", "", "User-entered label")
    add(T, "OLDKEY", "", "Legacy key (migration reference)", "", "For backwards compatibility")
    add(T, "FNAME", "", "Source filename", "", "Original data file reference")
    add(T, "STARTDATE", "", "Load profile start date", "", "Date string for load profile start")
    add(T, "BASEAMPS", "A", "Base ampacity / peak current", "", "Reference current for per-unit load shapes")
    add(T, "TAG", "", "Record tag", "", "Internal flag")
    add(T, "AVERAGE_LOADFACTOR", "", "Average load factor (calculated)", "", "Mean of daily load factors across profile")
    add(T, "NDAYS", "", "Number of days in profile", "", "Count of daily shapes linked")
    add(T, "PRELOAD", "", "Pre-loading flag or value", "", "For cyclic pre-loading before transient")

    # ===================================================================
    # SHAPEWIN (56 columns) — T/I pairs + metadata
    # ===================================================================
    T = "SHAPEWIN"
    add(T, "KEYSHAPE", "", "Shape ID — unique identifier", "", "Primary key, VARCHAR")
    add(T, "TITLE", "", "Shape title", "", "User-entered label")
    add(T, "SHAPETYPE", "", "Shape type code", "", "Type of daily load shape")
    add(T, "LOADFACT", "", "Load factor for this shape", "", "Average/peak ratio")
    add(T, "TAG", "", "Record tag", "", "Internal flag")
    add(T, "DATA_PAGE", "", "Data page reference (GUI)", "", "Which GUI tab the data is on")
    # T0/I0 through T23/I23 = 48 columns
    for h in range(24):
        add(T, f"T{h}", "h", f"Time point {h} (hour offset from midnight)", "", "24-hour load shape time coordinate")
        add(T, f"I{h}", "p.u.", f"Load at time point {h} (per unit of base amps)", "", "24-hour load shape current value")

    # ===================================================================
    # DAYWIN (5 columns) — fully documented
    # ===================================================================
    T = "DAYWIN"
    add(T, "KEYLOAD", "", "Load window ID — foreign key to LOADWIN", "", "Links day to parent load window")
    add(T, "KEYSHAPE", "", "Shape ID — foreign key to SHAPEWIN", "", "Which daily shape to use")
    add(T, "SF", "", "Scale factor for this day", "", "Multiplier on base amps for this day")
    add(T, "ORDER_NO", "", "Day order number within load window", "", "Sequence position")
    add(T, "TAG", "", "Record tag", "", "Internal flag")

    # ===================================================================
    # DUCTLIB (33 columns) — duct bank geometry library
    # ===================================================================
    T = "DUCTLIB"
    add(T, "DUCTBANKID", "", "Duct bank ID — unique identifier", "", "Primary key, VARCHAR")
    add(T, "DUCTBANK", "", "Duct bank name/title", "", "User-entered label")
    add(T, "TAG", "", "Record tag", "", "Internal flag")
    add(T, "HEIGHT", "m", "Duct bank overall height", "", "Vertical extent of encasement")
    add(T, "WIDTH", "m", "Duct bank overall width", "", "Horizontal extent of encasement")
    add(T, "AXB", "", "Duct bank drawing axis parameter (GUI)", "", "Cross-section display")
    add(T, "ALB", "", "Duct bank drawing scale parameter (GUI)", "", "Cross-section display")
    add(T, "OFFSETX", "m", "X offset of duct array from bank centre", "", "Duct positioning")
    add(T, "OFFSETY", "m", "Y offset of duct array from bank centre", "", "Duct positioning")
    add(T, "OFFSETX2", "m", "X offset (secondary array)", "", "For multi-array banks")
    add(T, "OFFSETY2", "m", "Y offset (secondary array)", "", "For multi-array banks")
    add(T, "DISTX", "m", "Horizontal spacing between duct centres", "", "Duct pitch X")
    add(T, "DISTY", "m", "Vertical spacing between duct centres", "", "Duct pitch Y")
    add(T, "TOTROW", "", "Total number of duct rows", "", "Bank layout")
    add(T, "TOTCOL", "", "Total number of duct columns", "", "Bank layout")
    add(T, "DUPIOTDI", "m", "Duct outside diameter", "", "External diameter of duct conduit")
    add(T, "DUPIINDI", "m", "Duct inside diameter", "", "Internal diameter of duct conduit")
    add(T, "XAXISORGM", "m", "X axis origin (metric, GUI)", "", "Drawing coordinate")
    add(T, "XAXISLENM", "m", "X axis length (metric, GUI)", "", "Drawing coordinate")
    add(T, "XAXISSTEPM", "m", "X axis step (metric, GUI)", "", "Drawing coordinate")
    add(T, "YAXISORGM", "m", "Y axis origin (metric, GUI)", "", "Drawing coordinate")
    add(T, "YAXISLENM", "m", "Y axis length (metric, GUI)", "", "Drawing coordinate")
    add(T, "YAXISSTEPM", "m", "Y axis step (metric, GUI)", "", "Drawing coordinate")
    add(T, "XAXISORGI", "in", "X axis origin (imperial, GUI)", "", "Drawing coordinate")
    add(T, "XAXISLENI", "in", "X axis length (imperial, GUI)", "", "Drawing coordinate")
    add(T, "XAXISSTEPI", "in", "X axis step (imperial, GUI)", "", "Drawing coordinate")
    add(T, "YAXISORGI", "in", "Y axis origin (imperial, GUI)", "", "Drawing coordinate")
    add(T, "YAXISLENI", "in", "Y axis length (imperial, GUI)", "", "Drawing coordinate")
    add(T, "YAXISSTEPI", "in", "Y axis step (imperial, GUI)", "", "Drawing coordinate")
    add(T, "LOCKEDBY", "", "Locked by user", "", "Multi-user locking field")
    add(T, "LOCKTSTAMP", "", "Lock timestamp", "", "Multi-user locking timestamp")
    add(T, "STD_NOMINAL_DUCT", "", "Standard nominal duct size designation", "", "Duct product standard")
    add(T, "CLEARANCES_MODE", "", "Clearances calculation mode", "", "How clearances are computed")

    # ===================================================================
    # MULTIPLELAYER (52 columns) — soil/backfill layered thermal model
    # ===================================================================
    T = "MULTIPLELAYER"
    add(T, "IDSTUDY", "", "Study ID — foreign key to STUDYHED", "", "Links to parent study")
    add(T, "EXECUNO", "", "Execution number", "", "Sub-case identifier")
    add(T, "TAG", "", "Record tag", "", "Internal flag")
    add(T, "LAYERNO", "", "Layer number (0-based)", "", "Index of this soil/backfill layer")
    add(T, "LAYERTYPE", "", "Layer type", "", "Soil, backfill, duct bank, etc.")
    add(T, "LAYERNAME", "", "Layer display name", "", "User-entered label")
    add(T, "LAYERCAPTION", "", "Layer caption (GUI)", "", "GUI display label")
    add(T, "XLEFT", "m", "Left edge X coordinate", "", "Layer boundary")
    add(T, "YTOP", "m", "Top edge Y coordinate (depth)", "", "Layer boundary")
    add(T, "XB", "m", "Layer width", "", "Horizontal extent")
    add(T, "YB", "m", "Layer height", "", "Vertical extent")
    add(T, "AXB", "", "Drawing axis parameter (GUI)", "", "Cross-section display")
    add(T, "AYB", "", "Drawing Y-axis parameter (GUI)", "", "Cross-section display")
    add(T, "DUCTYPOS", "m", "Duct bank Y position within layer", "", "Depth of duct bank")
    add(T, "DUCTBANKID", "", "Duct bank library ID", "", "FK to DUCTLIB")
    add(T, "DUCTBANK", "", "Duct bank name", "", "Display name")
    add(T, "RHD", "K.m/W", "Duct thermal resistivity", "", "Duct material property")
    add(T, "ALIGN_HOR", "", "Horizontal alignment (GUI)", "", "GUI display")
    add(T, "ALIGN_VER", "", "Vertical alignment (GUI)", "", "GUI display")
    add(T, "LAYERCOLOR", "", "Layer colour code (GUI)", "", "GUI display")
    add(T, "LAYERRESIS", "K.m/W", "Layer thermal resistivity", "", "Soil/backfill thermal resistivity")
    add(T, "INSIDE_DIAM", "m", "Casing inside diameter", "", "For casing within layer")
    add(T, "OUTSIDE_DIAM", "m", "Casing outside diameter", "", "For casing within layer")
    add(T, "CASING_CONFIGURATION_TYPE", "", "Casing configuration type", "", "Casing arrangement")
    add(T, "CASING_CONSTRUCTION_TYPE", "", "Casing construction type", "", "Casing material/build")
    add(T, "DUCT_ATTACHED_TO_CASING_INDEX", "", "Index of duct attached to casing", "", "Casing-duct linkage")
    add(T, "SHAPETYPE", "", "Layer shape type", "", "Cross-section shape (rectangular, trapezoidal, etc.)")
    add(T, "FRAME_XCENTRE", "m", "Frame X centre (GUI)", "", "GUI frame position")
    add(T, "FRAME_YCENTRE", "m", "Frame Y centre (GUI)", "", "GUI frame position")
    add(T, "FRAME_XLEFT", "m", "Frame left edge (GUI)", "", "GUI frame position")
    add(T, "FRAME_YTOP", "m", "Frame top edge (GUI)", "", "GUI frame position")
    add(T, "FRAME_WIDTH", "m", "Frame width (GUI)", "", "GUI frame size")
    add(T, "FRAME_HEIGHT", "m", "Frame height (GUI)", "", "GUI frame size")
    add(T, "FRAME_LAYERCOLOR", "", "Frame layer colour (GUI)", "", "GUI display")
    add(T, "FRAME_LAYERRESIS", "K.m/W", "Frame layer thermal resistivity", "", "Encasement material property")
    add(T, "DUCTXPOS", "m", "Duct X position within layer", "", "Horizontal position of duct")
    add(T, "LAYER_RESIS_SAME_AS_SOIL_RESIS_FLAG", "", "Layer resistivity same as soil flag", "", "Whether layer inherits soil properties")
    add(T, "DUCT_CUSTOM_U", "", "Custom duct thermal parameter U", "", "Custom thermal coefficient")
    add(T, "DUCT_CUSTOM_V", "", "Custom duct thermal parameter V", "", "Custom thermal coefficient")
    add(T, "DUCT_CUSTOM_Y", "", "Custom duct thermal parameter Y", "", "Custom thermal coefficient")
    add(T, "LD_ARROW_WIDTH_POSITION", "", "Layer dimension arrow width position (GUI)", "", "GUI dimension display")
    add(T, "LD_ARROW_WIDTH_MARGIN", "", "Layer dimension arrow width margin (GUI)", "", "GUI dimension display")
    add(T, "LD_ARROW_WIDTH_INSIDECOMPO_FLAG", "", "Width arrow inside component flag (GUI)", "", "GUI dimension display")
    add(T, "LD_ARROW_HEIGHT_POSITION", "", "Layer dimension arrow height position (GUI)", "", "GUI dimension display")
    add(T, "LD_ARROW_HEIGHT_MARGIN", "", "Layer dimension arrow height margin (GUI)", "", "GUI dimension display")
    add(T, "LD_ARROW_HEIGHT_INSIDECOMPO_FLAG", "", "Height arrow inside component flag (GUI)", "", "GUI dimension display")
    add(T, "MEDIUM_DUCT_FLAG", "", "Medium duct flag", "", "Duct fill medium type")
    add(T, "MAGNETIC_PROPERTY_FLAG", "", "Magnetic property flag", "", "Include magnetic effects")
    add(T, "FRAME_LAYERRESIS_NEW", "K.m/W", "Updated frame layer thermal resistivity", "", "Updated encasement property")
    add(T, "REAL_DUCTS_IN_TREFOIL_YCENTRE", "m", "Real duct trefoil Y centre", "", "Actual trefoil position")
    add(T, "DUCT_FILLING_TH_RESIS", "K.m/W", "Duct filling thermal resistivity", "", "Thermal resistivity of duct fill material")
    add(T, "DUCT_IN_TREFOIL_RECALCULATED_FLAG", "", "Duct trefoil recalculated flag", "", "Whether trefoil positions were recalculated")

    # ===================================================================
    # TRANSIEN (160 columns) — transient analysis configuration
    # ===================================================================
    T = "TRANSIEN"
    add(T, "IDSTUDY", "", "Study ID — foreign key to STUDYHED", "", "Links to parent study")
    add(T, "EXECUNO", "", "Execution number", "", "Sub-case identifier")
    add(T, "TAG", "", "Record tag", "", "Internal flag")
    add(T, "TRANSELCT", "", "Transient selection type", "", "Which transient analysis to run")
    add(T, "TEMP", "°C", "Target/initial temperature", "", "Starting or target conductor temperature")
    add(T, "TIMEREQ", "h", "Time required", "", "Duration of transient event")
    add(T, "LCURV", "", "Load curve type", "", "Shape of transient load profile")
    add(T, "SRCH1", "", "Search parameter 1", "", "Solver search control")
    add(T, "SRCH2", "", "Search parameter 2", "", "Solver search control")
    add(T, "RANGT1", "", "Range parameter T1", "", "Temperature/time range bound")
    add(T, "RANGT2", "", "Range parameter T2", "", "Temperature/time range bound")
    add(T, "SCALE1", "", "Scale parameter 1", "", "Solver scaling")
    add(T, "SCALE2", "", "Scale parameter 2", "", "Solver scaling")
    add(T, "RESSC1", "", "Result scaling parameter", "", "Output scaling")
    add(T, "REST1", "", "Result time parameter", "", "Output time range")
    add(T, "SOLVTYPOP1", "", "Solver type option", "", "Solver algorithm selection")
    add(T, "PARTICDATA", "", "Particular data (serialized)", "", "Additional transient parameters")
    add(T, "SCFACTDATA", "", "Scale factor data (serialized)", "", "Time-dependent scale factors")
    add(T, "DESTEMDATA", "", "Design temperature data (serialized)", "", "Temperature constraints")
    add(T, "COMTYP", "", "Computation type", "", "Detailed calculation method")
    # IDLOAD1-45: Load window references for up to 45 cables
    for i in range(1, 46):
        add(T, f"IDLOAD{i}", "", f"Load window ID for cable {i}", "", "FK to LOADWIN — assigns load profile to cable position")
    # FACT1-45: Scale factors for each cable
    for i in range(1, 46):
        add(T, f"FACT{i}", "", f"Load scale factor for cable {i}", "", "Multiplier on load profile for this cable")
    # TOG1-45: Toggle flags for each cable
    for i in range(1, 46):
        add(T, f"TOG{i}", "", f"Toggle flag for cable {i}", "", "Enable/disable this cable in transient calc")
    add(T, "EMERCURDATA", "", "Emergency current data (serialized)", "", "Emergency overload current profile")
    add(T, "DES_SHEATHTEMDATA", "", "Design sheath temperature data (serialized)", "", "Sheath temperature constraints")
    add(T, "DES_ARMOURTEMDATA", "", "Design armour temperature data (serialized)", "", "Armour temperature constraints")
    add(T, "DES_EXTERIORTEMDATA", "", "Design exterior temperature data (serialized)", "", "External surface temperature constraints")
    add(T, "DATAPAGE", "", "Data page reference (GUI)", "", "Which GUI tab the data is on")

    # ===================================================================
    # STUDCAB (233 columns) — mirrors CABLES + study keys
    # ===================================================================
    T = "STUDCAB"
    add(T, "IDSTUDY", "", "Study ID — foreign key to STUDYHED", "", "Links to parent study")
    add(T, "EXECUNO", "", "Execution number", "", "Sub-case identifier")
    # All other columns mirror CABLES table — handled by fallback to CABLES knowledge

    # ===================================================================
    # STUDCAB_LAYERS (13 columns) — mirrors CABLE_LAYERS + study keys
    # ===================================================================
    T = "STUDCAB_LAYERS"
    add(T, "IDSTUDY", "", "Study ID — foreign key to STUDYHED", "", "Links to parent study")
    add(T, "EXECUNO", "", "Execution number", "", "Sub-case identifier")
    # Rest mirrors CABLE_LAYERS — handled by fallback

    # ===================================================================
    # VERSION (1 column)
    # ===================================================================
    add("VERSION", "VERSION", "", "CYMCAP database schema version string", "", "Tracks DB format version")

    # ===================================================================
    # CALCTYPE (3 columns)
    # ===================================================================
    T = "CALCTYPE"
    add(T, "IDSTUDY", "", "Study ID — foreign key to STUDYHED", "", "Links to parent study")
    add(T, "EXECUNO", "", "Execution number", "", "Sub-case identifier")
    add(T, "CALCTYPE", "", "Calculation type code", "", "Which calculation method was used")

    # ===================================================================
    # HEATSRCLIB (6 columns)
    # ===================================================================
    T = "HEATSRCLIB"
    add(T, "HEATSOURCE_ID", "", "Heat source ID — unique identifier", "", "Primary key")
    add(T, "TITLE", "", "Heat source title", "", "User-entered label")
    add(T, "TAG", "", "Record tag", "", "Internal flag")
    add(T, "HEAT_RATE", "W/m", "Heat dissipation rate per unit length", "", "External heat source")
    add(T, "OUTSIDE_DIAM", "m", "Outside diameter of heat source", "", "Physical size")
    add(T, "HEATSOURCE_TYPE", "", "Heat source type code", "", "Type of external heat source")

    # ===================================================================
    # ALSSS (24 columns) — ampacity/loss/steady-state results
    # ===================================================================
    T = "ALSSS"
    add(T, "IDSTUDY", "", "Study ID — foreign key to STUDYHED", "", "Links to parent study")
    add(T, "EXECUNO", "", "Execution number", "", "Sub-case identifier")
    add(T, "IDCAB", "", "Cable ID — foreign key to CABLES", "", "Which cable these results are for")
    add(T, "TAG", "", "Record tag", "", "Internal flag")
    for i in range(1, 11):
        add(T, f"ALS{i}", "A", f"Ampacity result {i}", "", "Rated current for calculation case")
        add(T, f"SS{i}", "°C", f"Steady-state temperature result {i}", "", "Maximum conductor temperature")

    # ===================================================================
    # GLOBAL_ENV_PARAM_AMBIENT_VS_DEPTH_RANGE (5 columns)
    # ===================================================================
    T = "GLOBAL_ENV_PARAM_AMBIENT_VS_DEPTH_RANGE"
    add(T, "ENVIRONMENT_CONFIG_ID", "", "Config ID — FK to GLOBAL_ENVIRONMENT_PARAMETERS", "", "Parent config")
    add(T, "DEPTH", "m", "Burial depth", "", "Depth below surface")
    add(T, "AMBIENT_TEMP", "°C", "Ambient temperature at this depth", "", "Temperature vs depth profile")
    add(T, "TAG", "", "Record tag", "", "Internal flag")
    add(T, "ORDER_NO", "", "Sequence order", "", "Position in depth profile")

    # ===================================================================
    # Historical tables (empty — mirror LOADWIN/SHAPEWIN/DAYWIN)
    # ===================================================================
    T = "HISTORICAL_LOADWIN"
    add(T, "KEYLOAD", "", "Load window ID", "", "Same as LOADWIN.KEYLOAD")
    add(T, "TITLE", "", "Load window title", "", "Same as LOADWIN")
    add(T, "OLDKEY", "", "Legacy key", "", "Same as LOADWIN")
    add(T, "FNAME", "", "Source filename", "", "Same as LOADWIN")
    add(T, "STARTDATE", "", "Start date", "", "Same as LOADWIN")
    add(T, "BASEAMPS", "A", "Base ampacity", "", "Same as LOADWIN")
    add(T, "TAG", "", "Record tag", "", "Internal flag")
    add(T, "AVERAGE_LOADFACTOR", "", "Average load factor", "", "Same as LOADWIN")
    add(T, "NDAYS", "", "Number of days", "", "Same as LOADWIN")
    add(T, "PRELOAD", "", "Pre-loading flag", "", "Same as LOADWIN")

    T = "HISTORICAL_SHAPEWIN"
    add(T, "KEYSHAPE", "", "Shape ID", "", "Same as SHAPEWIN.KEYSHAPE")
    add(T, "TITLE", "", "Shape title", "", "Same as SHAPEWIN")
    add(T, "SHAPETYPE", "", "Shape type code", "", "Same as SHAPEWIN")
    add(T, "LOADFACT", "", "Load factor", "", "Same as SHAPEWIN")
    add(T, "TAG", "", "Record tag", "", "Internal flag")
    add(T, "DATA_PAGE", "", "Data page reference", "", "Same as SHAPEWIN")
    for h in range(24):
        add(T, f"T{h}", "h", f"Time point {h}", "", "Same as SHAPEWIN")
        add(T, f"I{h}", "p.u.", f"Load at time point {h}", "", "Same as SHAPEWIN")

    T = "HISTORICAL_DAYWIN"
    add(T, "KEYLOAD", "", "Load window ID", "", "Same as DAYWIN")
    add(T, "KEYSHAPE", "", "Shape ID", "", "Same as DAYWIN")
    add(T, "SF", "", "Scale factor", "", "Same as DAYWIN")
    add(T, "ORDER_NO", "", "Day order number", "", "Same as DAYWIN")
    add(T, "TAG", "", "Record tag", "", "Internal flag")
    add(T, "HISTORICAL_ORDER", "", "Historical ordering index", "", "Position in historical sequence")

    # ===================================================================
    # JOBSTEMPLATE (184 columns) — mirrors STUDYHED
    # ===================================================================
    # Same columns as STUDYHED — handled by fallback

    # ===================================================================
    # Small/empty tables
    # ===================================================================
    T = "CASINGLIB"
    add(T, "CASINGID", "", "Casing ID — unique identifier", "", "Primary key")
    add(T, "TITLE", "", "Casing title", "", "User-entered label")
    add(T, "TAG", "", "Record tag", "", "Internal flag")
    add(T, "CASINGMAT", "", "Casing material code", "", "Material of casing pipe")
    add(T, "INSIDEDIAM", "m", "Inside diameter", "", "Casing bore")
    add(T, "OUTSIDEDIAM", "m", "Outside diameter", "", "Casing OD")
    add(T, "RH", "K.m/W", "Thermal resistivity", "", "Casing material thermal resistivity")
    add(T, "CASING_NAME", "", "Casing name", "", "User-entered casing label")
    add(T, "XCENTRE", "m", "Casing centre X coordinate", "", "Horizontal position")
    add(T, "YCENTRE", "m", "Casing centre Y coordinate (depth)", "", "Vertical position")
    add(T, "NBDUCTS", "", "Number of ducts in casing", "", "Duct count within casing")
    add(T, "DISTRAD", "m", "Radial distance of ducts from casing centre", "", "Duct positioning radius")
    add(T, "STARTANGLE", "°", "Starting angle for duct arrangement", "", "Angular offset for first duct")
    add(T, "DUCTS_TOUCHING", "", "Ducts touching flag", "", "Whether ducts are in contact")
    add(T, "CENTRE_DUCT", "", "Centre duct flag", "", "Whether there is a centre duct")
    add(T, "DUPIOTDI", "m", "Duct outside diameter", "", "External diameter of duct conduit")
    add(T, "DUPIINDI", "m", "Duct inside diameter", "", "Internal diameter of duct conduit")
    add(T, "CASINGINDI", "m", "Casing inside diameter (alternative)", "", "Casing bore dimension")
    add(T, "CASINGOTDI", "m", "Casing outside diameter (alternative)", "", "Casing OD dimension")
    add(T, "STD_NOMINAL_DUCT", "", "Standard nominal duct size", "", "Duct product standard")
    add(T, "LOCKEDBY", "", "Locked by user", "", "Multi-user locking field")
    add(T, "LOCKTSTAMP", "", "Lock timestamp", "", "Multi-user locking timestamp")

    T = "MCPP"
    add(T, "IDSTUDY", "", "Study ID — FK to STUDYHED", "", "Parent study")
    add(T, "EXECUNO", "", "Execution number", "", "Sub-case")
    add(T, "TAG", "", "Record tag", "", "Internal flag")
    add(T, "CIRCUITNO", "", "Circuit number", "", "Which circuit for multi-cable per phase")
    add(T, "ROWNO", "", "Row number", "", "Data row index")
    add(T, "CURRENT_PHASEA", "A", "Phase A current", "", "Current in phase A")
    add(T, "CURRENT_PHASEB", "A", "Phase B current", "", "Current in phase B")
    add(T, "CURRENT_PHASEC", "A", "Phase C current", "", "Current in phase C")

    T = "MULTIPLE_HEATSOURCE"
    add(T, "IDSTUDY", "", "Study ID — FK to STUDYHED", "", "Parent study")
    add(T, "EXECUNO", "", "Execution number", "", "Sub-case")
    add(T, "TAG", "", "Record tag", "", "Internal flag")
    add(T, "HEATSOURCE_ID", "", "Heat source ID — FK to HEATSRCLIB", "", "Which heat source")
    add(T, "X", "m", "Horizontal position", "", "X coordinate")
    add(T, "Y", "m", "Vertical position (depth)", "", "Y coordinate")
    add(T, "HEATSRCNO", "", "Heat source number", "", "Index of heat source within study")
    add(T, "HEORTE", "", "Heat or temperature boundary type", "", "Whether source is heat rate or fixed temperature")
    add(T, "HEATEMP", "°C", "Heat source temperature", "", "Fixed temperature if temperature boundary")
    add(T, "XLHEAT", "m", "Heat source X position (label)", "", "Label X coordinate")
    add(T, "YLHEAT", "m", "Heat source Y position (label)", "", "Label Y coordinate")
    add(T, "DIAHEAT", "m", "Heat source diameter", "", "Physical size of heat source")
    add(T, "HEATSOURCE_INSULATION_FLAG", "", "Heat source insulation flag", "", "Whether heat source has insulation")
    add(T, "INSUDIAM_HEATSOURCE", "m", "Heat source insulation diameter", "", "Insulation OD of heat source")
    add(T, "THRESIS_HEATSOURCE", "K.m/W", "Heat source insulation thermal resistivity", "", "Insulation of heat source")
    add(T, "HEATSOURCE_COLD_OR_HOT", "", "Heat source hot/cold flag", "", "Whether source heats or cools")
    add(T, "HEATSRC_ID", "", "Heat source library ID", "", "FK to HEATSRCLIB")
    add(T, "HEATSRC_SF", "", "Heat source scale factor", "", "Multiplier on library heat rate")

    T = "PS_SCENARIO"
    add(T, "IDSTUDY", "", "Study ID — FK to STUDYHED", "", "Parent study")
    add(T, "EXECUNO", "", "Execution number", "", "Sub-case")
    add(T, "SCENARIO_NO", "", "Scenario sequence number", "", "Ordering index")
    add(T, "TITLE", "", "Scenario title", "", "Label")
    add(T, "TAG", "", "Record tag", "", "Internal flag")

    T = "PS_SPECVALUES"
    add(T, "IDSTUDY", "", "Study ID — FK to STUDYHED", "", "Parent study")
    add(T, "EXECUNO", "", "Execution number", "", "Sub-case")
    add(T, "SCENARIO_NO", "", "Scenario sequence number", "", "FK to PS_SCENARIO")
    add(T, "VARIABLE_NO", "", "Variable sequence number", "", "FK to PS_VARIABLE")
    add(T, "COLUMN_NO", "", "Column number in parameter grid", "", "Grid position")
    add(T, "ROW_NO", "", "Row number in parameter grid", "", "Grid position")
    add(T, "KEY_VAL1", "", "Key value 1 (parameter identifier)", "", "First key for the parameter")
    add(T, "KEY_VAL2", "", "Key value 2 (parameter identifier)", "", "Second key for the parameter")
    add(T, "VALUE1", "", "Parameter value 1", "", "First value for this scenario/variable")
    add(T, "VALUE2", "", "Parameter value 2", "", "Second value for this scenario/variable")
    add(T, "TAG", "", "Record tag", "", "Internal flag")

    T = "PS_VARIABLE"
    add(T, "IDSTUDY", "", "Study ID — FK to STUDYHED", "", "Parent study")
    add(T, "EXECUNO", "", "Execution number", "", "Sub-case")
    add(T, "SCENARIO_NO", "", "Scenario sequence number", "", "FK to PS_SCENARIO")
    add(T, "VARIABLE_NO", "", "Variable sequence number", "", "Ordering index")
    add(T, "TITLE", "", "Variable title", "", "Label")
    add(T, "TAG", "", "Record tag", "", "Internal flag")
    add(T, "VARIABLE_ID", "", "Variable identifier string", "", "User-visible variable ID")
    add(T, "CATEGORY_ID", "", "Variable category ID", "", "Grouping category")
    add(T, "COLUMN_NO", "", "Column number in parameter grid", "", "Grid position")
    add(T, "INPUT_TYPE", "", "Input type code", "", "How the variable is entered")
    add(T, "KEY_VAL1", "", "Key value 1 (parameter identifier)", "", "First key")
    add(T, "KEY_VAL2", "", "Key value 2 (parameter identifier)", "", "Second key")
    add(T, "FROM_VALUE", "", "Range start value", "", "Sweep start")
    add(T, "TO_VALUE", "", "Range end value", "", "Sweep end")
    add(T, "STEP_VALUE", "", "Range step value", "", "Sweep increment")

    T = "RTTR"
    add(T, "IDSTUDY", "", "Study ID — FK to STUDYHED", "", "Real-time thermal rating config")
    add(T, "EXECUNO", "", "Execution number", "", "Sub-case")
    add(T, "TAG", "", "Record tag", "", "Internal flag")
    add(T, "SECTION_FOLDER", "", "RTTR section/folder reference", "", "Data source location")
    add(T, "CALCULATION_TYPE", "", "RTTR calculation type", "", "Real-time calculation method")
    add(T, "FIBER_LOCATION", "", "Fibre optic location reference", "", "DTS fibre sensor position")
    add(T, "MAXIMUM_SOLVING_TIME_IN_SECONDS", "s", "Maximum solving time", "", "Solver timeout")
    add(T, "CIRCUITNO", "", "Circuit number", "", "Which circuit for RTTR")
    add(T, "ROWNO", "", "Row number", "", "Data row index")
    add(T, "CURRENT_PHASEA", "A", "Phase A current", "", "Measured phase A current")
    add(T, "CURRENT_PHASEB", "A", "Phase B current", "", "Measured phase B current")
    add(T, "CURRENT_PHASEC", "A", "Phase C current", "", "Measured phase C current")

    T = "UNDELETE"
    add(T, "ID", "", "Unique record identifier", "", "Auto-increment ID")
    add(T, "KEY", "", "Primary key of deleted record", "", "For undo functionality")
    add(T, "KEY2", "", "Secondary key of deleted record", "", "For composite-key tables")
    add(T, "DATE", "", "Deletion date", "", "When the record was deleted")
    add(T, "TITLE", "", "Deleted record title/description", "", "Summary of what was deleted")

    # ===================================================================
    # Additional fields for tables with stragglers
    # ===================================================================
    add("CALCTYPE", "TAG", "", "Record tag", "", "Internal flag")

    T = "GLOBAL_ENV_PARAM_AMBIENT_VS_DEPTH_RANGE"
    add(T, "ENVIRONMENT_CONFIG_NAME", "", "Configuration name", "", "Denormalized from parent config")
    add(T, "DEPTH_RANGE", "", "Depth range category", "", "Classification of depth band")

    T = "HEATSRCLIB"
    add(T, "HEATSRC_ID", "", "Heat source library ID (alternative key)", "", "Alternative identifier")
    add(T, "HEORTE", "", "Heat or temperature boundary type", "", "Whether source is heat rate or fixed temperature")
    add(T, "LOCKEDBY", "", "Locked by user", "", "Multi-user locking field")
    add(T, "LOCKTSTAMP", "", "Lock timestamp", "", "Multi-user locking timestamp")

    # Additional LOADWIN/SHAPEWIN locking fields
    add("LOADWIN", "LOCKEDBY", "", "Locked by user", "", "Multi-user locking field")
    add("LOADWIN", "LOCKTSTAMP", "", "Lock timestamp", "", "Multi-user locking timestamp")
    add("SHAPEWIN", "LOCKEDBY", "", "Locked by user", "", "Multi-user locking field")
    add("SHAPEWIN", "LOCKTSTAMP", "", "Lock timestamp", "", "Multi-user locking timestamp")

    # Historical extras
    add("HISTORICAL_LOADWIN", "HISTORICAL_DATA_FLAG", "", "Historical data flag", "", "Whether this is archived historical data")
    add("HISTORICAL_LOADWIN", "LOCKEDBY", "", "Locked by user", "", "Multi-user locking field")
    add("HISTORICAL_LOADWIN", "LOCKTSTAMP", "", "Lock timestamp", "", "Multi-user locking timestamp")
    add("HISTORICAL_SHAPEWIN", "PEAKAMPS", "A", "Peak amps for this shape", "", "Historical peak current")
    add("HISTORICAL_SHAPEWIN", "HISTORICAL_DATA_FLAG", "", "Historical data flag", "", "Whether this is archived historical data")
    add("HISTORICAL_SHAPEWIN", "LOCKEDBY", "", "Locked by user", "", "Multi-user locking field")
    add("HISTORICAL_SHAPEWIN", "LOCKTSTAMP", "", "Lock timestamp", "", "Multi-user locking timestamp")
    add("HISTORICAL_DAYWIN", "HISTORICAL_DATA_FLAG", "", "Historical data flag", "", "Whether this is archived historical data")

    # STUDCAB extras (beyond CABLES mirror)
    add("STUDCAB", "CrossBondedTransposition", "", "Cross-bonded transposition configuration", "", "Bonding arrangement for study cable")
    add("STUDCAB", "DUCT_FILLING_TH_RESIS", "K.m/W", "Duct filling thermal resistivity", "", "Thermal resistivity of duct fill material")

    return k


# Fallback tables: STUDCAB mirrors CABLES, STUDCAB_LAYERS mirrors CABLE_LAYERS,
# JOBSTEMPLATE mirrors STUDYHED, MULTIPLELAYER_JOBSTEMPLATE mirrors MULTIPLELAYER
MIRROR_TABLES = {
    "STUDCAB": "CABLES",
    "STUDCAB_LAYERS": "CABLE_LAYERS",
    "JOBSTEMPLATE": "STUDYHED",
    "MULTIPLELAYER_JOBSTEMPLATE": "MULTIPLELAYER",
    "HISTORICAL_LOADWIN": "LOADWIN",
    "HISTORICAL_SHAPEWIN": "SHAPEWIN",
    "HISTORICAL_DAYWIN": "DAYWIN",
}


def _lookup(knowledge: dict, table: str, col: str) -> dict:
    """Look up column metadata with fallback to mirrored tables."""
    key = (table.upper(), col.upper())
    if key in knowledge:
        return knowledge[key]

    # Try mirror table
    mirror = MIRROR_TABLES.get(table.upper())
    if mirror:
        mirror_key = (mirror, col.upper())
        if mirror_key in knowledge:
            info = knowledge[mirror_key].copy()
            info["justification"] = f"Same as {mirror}.{col}; " + info.get("justification", "")
            return info

    return {
        "units": "",
        "description": "Unknown — not yet documented",
        "enum_values": "",
        "justification": "Field not documented in available sources",
    }


# ---------------------------------------------------------------------------
# Database helpers
# ---------------------------------------------------------------------------

def get_tables(conn: pyodbc.Connection) -> list[tuple[str, int]]:
    """Return (table_name, row_count) sorted: data tables first, then empty."""
    cursor = conn.cursor()
    # Collect all table names first — cursor.execute() disrupts the tables() iterator
    table_names = [
        row.table_name
        for row in cursor.tables(tableType="TABLE")
        if not row.table_name.startswith("MSys")
    ]
    tables = []
    for name in table_names:
        cursor.execute(f"SELECT COUNT(*) FROM [{name}]")
        count = cursor.fetchone()[0]
        tables.append((name, count))
    # Sort: data-bearing first (by descending row count), then empty (alphabetical)
    data_tables = sorted([(n, c) for n, c in tables if c > 0], key=lambda x: -x[1])
    empty_tables = sorted([(n, c) for n, c in tables if c == 0])
    return data_tables + empty_tables


def get_columns(conn: pyodbc.Connection, table: str) -> list[dict[str, str]]:
    """Return list of {name, type_name} for a table."""
    cursor = conn.cursor()
    return [
        {"name": row.column_name, "type_name": row.type_name}
        for row in cursor.columns(table=table)
    ]


# ---------------------------------------------------------------------------
# Excel generation
# ---------------------------------------------------------------------------

HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGN = Alignment(vertical="top", wrap_text=True)


def write_table_sheet(ws, table_name: str, row_count: int, columns: list[dict], knowledge: dict) -> None:
    """Write one table's data dictionary to a worksheet."""
    headers = ["Column Name", "DB Type", "Assumed Units", "Description", "Enum Values", "Justification"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    for row_idx, col_info in enumerate(columns, 2):
        col_name = col_info["name"]
        db_type = col_info["type_name"]
        meta = _lookup(knowledge, table_name, col_name)

        ws.cell(row=row_idx, column=1, value=col_name).alignment = CELL_ALIGN
        ws.cell(row=row_idx, column=2, value=db_type).alignment = CELL_ALIGN
        ws.cell(row=row_idx, column=3, value=meta["units"]).alignment = CELL_ALIGN
        ws.cell(row=row_idx, column=4, value=meta["description"]).alignment = CELL_ALIGN
        ws.cell(row=row_idx, column=5, value=meta["enum_values"]).alignment = CELL_ALIGN
        ws.cell(row=row_idx, column=6, value=meta["justification"]).alignment = CELL_ALIGN

    # Auto-width columns (approximate)
    col_widths = [20, 12, 12, 50, 40, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze header row
    ws.freeze_panes = "A2"


def main() -> None:
    if not DB_PATH.exists():
        print(f"ERROR: Database not found: {DB_PATH}", file=sys.stderr)
        sys.exit(1)

    print(f"Connecting to {DB_PATH.name}...")
    conn_str = f"DRIVER={{Microsoft Access Driver (*.mdb, *.accdb)}};DBQ={DB_PATH};"
    conn = pyodbc.connect(conn_str)

    print("Building knowledge dictionary...")
    knowledge = _build_knowledge_dict()

    print("Reading database schema...")
    tables = get_tables(conn)

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for table_name, row_count in tables:
        print(f"  {table_name}: {row_count} rows...")
        columns = get_columns(conn, table_name)

        # Sheet name max 31 chars
        sheet_name = table_name[:31]
        ws = wb.create_sheet(title=sheet_name)
        write_table_sheet(ws, table_name, row_count, columns, knowledge)

    conn.close()

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"\nSaved: {OUTPUT_PATH}")
    print(f"Total sheets: {len(wb.sheetnames)}")

    # Quick verification
    documented = 0
    total = 0
    conn2 = pyodbc.connect(conn_str)
    for table_name, _ in tables:
        columns = get_columns(conn2, table_name)
        for col in columns:
            total += 1
            meta = _lookup(knowledge, table_name, col["name"])
            if meta["description"] != "Unknown — not yet documented":
                documented += 1
    conn2.close()
    print(f"Documented columns: {documented}/{total} ({100*documented/total:.0f}%)")


if __name__ == "__main__":
    main()
