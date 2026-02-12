#!/usr/bin/env python3
"""
Excel MCP Server
================
An MCP server that allows Claude to analyze Excel files including:
- Cell formulas
- VBA macros
- Named ranges
- Formula dependencies
- Sheet structure

Author: Hrit Kandel
"""

import json
import re
import sys
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

# MCP SDK imports
from mcp.server import Server
from mcp.server.stdio import stdio_server
from mcp.types import (
    Tool,
    TextContent,
    Resource,
    ResourceTemplate,
)

# Optional: VBA extraction
try:
    from oletools import olevba
    HAS_OLETOOLS = True
except ImportError:
    HAS_OLETOOLS = False

# Initialize MCP server
server = Server("excel-analyzer")


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def parse_formula_references(formula: str) -> list[str]:
    """Extract cell references from a formula."""
    # Match cell references like A1, $A$1, Sheet1!A1, 'Sheet Name'!A1
    pattern = r"(?:'[^']+'\!|\w+\!)?\$?[A-Z]+\$?\d+"
    matches = re.findall(pattern, formula, re.IGNORECASE)
    return list(set(matches))


def get_formula_dependencies(wb: openpyxl.Workbook, sheet_name: str, cell_ref: str) -> dict:
    """Analyze what cells a formula depends on."""
    sheet = wb[sheet_name]
    cell = sheet[cell_ref]
    
    if not cell.value or not isinstance(cell.value, str) or not cell.value.startswith('='):
        return {"cell": cell_ref, "is_formula": False, "value": cell.value}
    
    refs = parse_formula_references(cell.value)
    return {
        "cell": cell_ref,
        "is_formula": True,
        "formula": cell.value,
        "references": refs,
        "reference_count": len(refs)
    }


def load_workbook_safe(filepath: str) -> tuple[openpyxl.Workbook | None, str | None]:
    """Safely load a workbook with error handling."""
    try:
        path = Path(filepath).expanduser().resolve()
        if not path.exists():
            return None, f"File not found: {filepath}"
        if not path.suffix.lower() in ['.xlsx', '.xlsm', '.xltx', '.xltm']:
            return None, f"Unsupported file type: {path.suffix}"
        
        wb = openpyxl.load_workbook(str(path), data_only=False, keep_vba=True)
        return wb, None
    except Exception as e:
        return None, f"Error loading workbook: {str(e)}"


# =============================================================================
# MCP TOOLS
# =============================================================================

@server.list_tools()
async def list_tools() -> list[Tool]:
    """List all available Excel analysis tools."""
    return [
        Tool(
            name="list_sheets",
            description="List all sheets in an Excel workbook",
            inputSchema={
                "type": "object",
                "properties": {
                    "filepath": {
                        "type": "string",
                        "description": "Path to the Excel file (.xlsx or .xlsm)"
                    }
                },
                "required": ["filepath"]
            }
        ),
        Tool(
            name="get_all_formulas",
            description="Extract all formulas from an Excel workbook, organized by sheet",
            inputSchema={
                "type": "object",
                "properties": {
                    "filepath": {
                        "type": "string",
                        "description": "Path to the Excel file"
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "Optional: specific sheet name (if omitted, returns all sheets)"
                    }
                },
                "required": ["filepath"]
            }
        ),
        Tool(
            name="get_cell_info",
            description="Get detailed information about a specific cell including formula, value, and dependencies",
            inputSchema={
                "type": "object",
                "properties": {
                    "filepath": {
                        "type": "string",
                        "description": "Path to the Excel file"
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "Name of the sheet"
                    },
                    "cell_ref": {
                        "type": "string",
                        "description": "Cell reference (e.g., 'A1', 'B5')"
                    }
                },
                "required": ["filepath", "sheet_name", "cell_ref"]
            }
        ),
        Tool(
            name="get_named_ranges",
            description="Get all named ranges/defined names in the workbook",
            inputSchema={
                "type": "object",
                "properties": {
                    "filepath": {
                        "type": "string",
                        "description": "Path to the Excel file"
                    }
                },
                "required": ["filepath"]
            }
        ),
        Tool(
            name="get_vba_code",
            description="Extract VBA macro code from a macro-enabled workbook (.xlsm)",
            inputSchema={
                "type": "object",
                "properties": {
                    "filepath": {
                        "type": "string",
                        "description": "Path to the Excel file (.xlsm)"
                    },
                    "module_name": {
                        "type": "string",
                        "description": "Optional: specific VBA module name"
                    }
                },
                "required": ["filepath"]
            }
        ),
        Tool(
            name="analyze_formula_chain",
            description="Trace the dependency chain for a formula - what cells it depends on and what depends on it",
            inputSchema={
                "type": "object",
                "properties": {
                    "filepath": {
                        "type": "string",
                        "description": "Path to the Excel file"
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "Name of the sheet"
                    },
                    "cell_ref": {
                        "type": "string",
                        "description": "Cell reference to analyze"
                    }
                },
                "required": ["filepath", "sheet_name", "cell_ref"]
            }
        ),
        Tool(
            name="get_sheet_structure",
            description="Get the structure of a sheet including used range, data regions, and summary",
            inputSchema={
                "type": "object",
                "properties": {
                    "filepath": {
                        "type": "string",
                        "description": "Path to the Excel file"
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "Name of the sheet"
                    }
                },
                "required": ["filepath", "sheet_name"]
            }
        ),
        Tool(
            name="search_formulas",
            description="Search for formulas containing specific functions or references",
            inputSchema={
                "type": "object",
                "properties": {
                    "filepath": {
                        "type": "string",
                        "description": "Path to the Excel file"
                    },
                    "search_term": {
                        "type": "string",
                        "description": "Function or text to search for (e.g., 'VLOOKUP', 'SUM', 'Sheet2!')"
                    }
                },
                "required": ["filepath", "search_term"]
            }
        ),
        Tool(
            name="get_full_analysis",
            description="Get a comprehensive analysis of the entire workbook including all formulas, named ranges, and VBA code",
            inputSchema={
                "type": "object",
                "properties": {
                    "filepath": {
                        "type": "string",
                        "description": "Path to the Excel file"
                    }
                },
                "required": ["filepath"]
            }
        )
    ]


@server.call_tool()
async def call_tool(name: str, arguments: dict[str, Any]) -> list[TextContent]:
    """Handle tool calls."""
    
    filepath = arguments.get("filepath", "")
    
    # -------------------------------------------------------------------------
    # LIST SHEETS
    # -------------------------------------------------------------------------
    if name == "list_sheets":
        wb, error = load_workbook_safe(filepath)
        if error:
            return [TextContent(type="text", text=f"Error: {error}")]
        
        sheets_info = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheets_info.append({
                "name": sheet_name,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "dimensions": sheet.dimensions
            })
        
        wb.close()
        return [TextContent(type="text", text=json.dumps({"sheets": sheets_info}, indent=2))]
    
    # -------------------------------------------------------------------------
    # GET ALL FORMULAS
    # -------------------------------------------------------------------------
    elif name == "get_all_formulas":
        wb, error = load_workbook_safe(filepath)
        if error:
            return [TextContent(type="text", text=f"Error: {error}")]
        
        target_sheet = arguments.get("sheet_name")
        sheets_to_check = [target_sheet] if target_sheet else wb.sheetnames
        
        result = {}
        for sheet_name in sheets_to_check:
            if sheet_name not in wb.sheetnames:
                continue
            
            sheet = wb[sheet_name]
            formulas = {}
            
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formulas[cell.coordinate] = {
                            "formula": cell.value,
                            "references": parse_formula_references(cell.value)
                        }
            
            if formulas:
                result[sheet_name] = {
                    "formula_count": len(formulas),
                    "formulas": formulas
                }
        
        wb.close()
        return [TextContent(type="text", text=json.dumps(result, indent=2))]
    
    # -------------------------------------------------------------------------
    # GET CELL INFO
    # -------------------------------------------------------------------------
    elif name == "get_cell_info":
        wb, error = load_workbook_safe(filepath)
        if error:
            return [TextContent(type="text", text=f"Error: {error}")]
        
        sheet_name = arguments.get("sheet_name")
        cell_ref = arguments.get("cell_ref", "").upper()
        
        if sheet_name not in wb.sheetnames:
            wb.close()
            return [TextContent(type="text", text=f"Error: Sheet '{sheet_name}' not found")]
        
        sheet = wb[sheet_name]
        cell = sheet[cell_ref]
        
        # Get both formula and calculated value
        wb_data = openpyxl.load_workbook(filepath, data_only=True)
        cell_value = wb_data[sheet_name][cell_ref].value
        wb_data.close()
        
        info = {
            "cell": cell_ref,
            "sheet": sheet_name,
            "formula": cell.value if isinstance(cell.value, str) and cell.value.startswith('=') else None,
            "value": cell_value,
            "data_type": cell.data_type,
            "number_format": cell.number_format,
            "is_formula": isinstance(cell.value, str) and cell.value.startswith('='),
            "comment": cell.comment.text if cell.comment else None
        }
        
        if info["is_formula"]:
            info["references"] = parse_formula_references(cell.value)
        
        wb.close()
        return [TextContent(type="text", text=json.dumps(info, indent=2))]
    
    # -------------------------------------------------------------------------
    # GET NAMED RANGES
    # -------------------------------------------------------------------------
    elif name == "get_named_ranges":
        wb, error = load_workbook_safe(filepath)
        if error:
            return [TextContent(type="text", text=f"Error: {error}")]
        
        named_ranges = {}
        for defined_name in wb.defined_names.definedName:
            named_ranges[defined_name.name] = {
                "refers_to": defined_name.attr_text,
                "scope": defined_name.localSheetId if defined_name.localSheetId is not None else "Workbook"
            }
        
        wb.close()
        return [TextContent(type="text", text=json.dumps({
            "count": len(named_ranges),
            "named_ranges": named_ranges
        }, indent=2))]
    
    # -------------------------------------------------------------------------
    # GET VBA CODE
    # -------------------------------------------------------------------------
    elif name == "get_vba_code":
        if not HAS_OLETOOLS:
            return [TextContent(type="text", text="Error: oletools not installed. Run: pip install oletools")]
        
        path = Path(filepath).expanduser().resolve()
        if not path.exists():
            return [TextContent(type="text", text=f"Error: File not found: {filepath}")]
        
        try:
            vba_parser = olevba.VBA_Parser(str(path))
            
            if not vba_parser.detect_vba_macros():
                vba_parser.close()
                return [TextContent(type="text", text="No VBA macros found in this workbook")]
            
            target_module = arguments.get("module_name")
            modules = {}
            
            for (filename, stream_path, vba_filename, vba_code) in vba_parser.extract_macros():
                if target_module and vba_filename != target_module:
                    continue
                modules[vba_filename] = {
                    "code": vba_code,
                    "line_count": len(vba_code.splitlines())
                }
            
            vba_parser.close()
            
            return [TextContent(type="text", text=json.dumps({
                "module_count": len(modules),
                "modules": modules
            }, indent=2))]
            
        except Exception as e:
            return [TextContent(type="text", text=f"Error extracting VBA: {str(e)}")]
    
    # -------------------------------------------------------------------------
    # ANALYZE FORMULA CHAIN
    # -------------------------------------------------------------------------
    elif name == "analyze_formula_chain":
        wb, error = load_workbook_safe(filepath)
        if error:
            return [TextContent(type="text", text=f"Error: {error}")]
        
        sheet_name = arguments.get("sheet_name")
        cell_ref = arguments.get("cell_ref", "").upper()
        
        if sheet_name not in wb.sheetnames:
            wb.close()
            return [TextContent(type="text", text=f"Error: Sheet '{sheet_name}' not found")]
        
        sheet = wb[sheet_name]
        cell = sheet[cell_ref]
        
        # Get dependencies (what this cell depends on)
        dependencies = get_formula_dependencies(wb, sheet_name, cell_ref)
        
        # Find dependents (what depends on this cell)
        dependents = []
        search_pattern = re.compile(rf'\b{cell_ref}\b', re.IGNORECASE)
        
        for row in sheet.iter_rows():
            for c in row:
                if c.value and isinstance(c.value, str) and c.value.startswith('='):
                    if search_pattern.search(c.value):
                        dependents.append({
                            "cell": c.coordinate,
                            "formula": c.value
                        })
        
        result = {
            "target_cell": cell_ref,
            "sheet": sheet_name,
            "dependencies": dependencies,
            "dependents": {
                "count": len(dependents),
                "cells": dependents
            }
        }
        
        wb.close()
        return [TextContent(type="text", text=json.dumps(result, indent=2))]
    
    # -------------------------------------------------------------------------
    # GET SHEET STRUCTURE
    # -------------------------------------------------------------------------
    elif name == "get_sheet_structure":
        wb, error = load_workbook_safe(filepath)
        if error:
            return [TextContent(type="text", text=f"Error: {error}")]
        
        sheet_name = arguments.get("sheet_name")
        
        if sheet_name not in wb.sheetnames:
            wb.close()
            return [TextContent(type="text", text=f"Error: Sheet '{sheet_name}' not found")]
        
        sheet = wb[sheet_name]
        
        # Count formulas, values, empty cells
        formula_count = 0
        value_count = 0
        
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value:
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_count += 1
                    else:
                        value_count += 1
        
        # Get merged cells
        merged = [str(rng) for rng in sheet.merged_cells.ranges]
        
        structure = {
            "sheet_name": sheet_name,
            "dimensions": sheet.dimensions,
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
            "formula_count": formula_count,
            "value_count": value_count,
            "merged_cells": merged,
            "has_autofilter": sheet.auto_filter.ref is not None if sheet.auto_filter else False
        }
        
        wb.close()
        return [TextContent(type="text", text=json.dumps(structure, indent=2))]
    
    # -------------------------------------------------------------------------
    # SEARCH FORMULAS
    # -------------------------------------------------------------------------
    elif name == "search_formulas":
        wb, error = load_workbook_safe(filepath)
        if error:
            return [TextContent(type="text", text=f"Error: {error}")]
        
        search_term = arguments.get("search_term", "").upper()
        matches = []
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        if search_term in cell.value.upper():
                            matches.append({
                                "sheet": sheet_name,
                                "cell": cell.coordinate,
                                "formula": cell.value
                            })
        
        wb.close()
        return [TextContent(type="text", text=json.dumps({
            "search_term": search_term,
            "match_count": len(matches),
            "matches": matches
        }, indent=2))]
    
    # -------------------------------------------------------------------------
    # GET FULL ANALYSIS
    # -------------------------------------------------------------------------
    elif name == "get_full_analysis":
        wb, error = load_workbook_safe(filepath)
        if error:
            return [TextContent(type="text", text=f"Error: {error}")]
        
        path = Path(filepath).expanduser().resolve()
        
        analysis = {
            "file": str(path),
            "file_type": path.suffix,
            "sheets": {},
            "named_ranges": {},
            "vba_modules": {},
            "summary": {
                "total_sheets": len(wb.sheetnames),
                "total_formulas": 0,
                "total_named_ranges": 0,
                "has_vba": False
            }
        }
        
        # Analyze each sheet
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            formulas = {}
            
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formulas[cell.coordinate] = cell.value
            
            analysis["sheets"][sheet_name] = {
                "dimensions": sheet.dimensions,
                "formula_count": len(formulas),
                "formulas": formulas
            }
            analysis["summary"]["total_formulas"] += len(formulas)
        
        # Named ranges
        for defined_name in wb.defined_names.definedName:
            analysis["named_ranges"][defined_name.name] = defined_name.attr_text
        analysis["summary"]["total_named_ranges"] = len(analysis["named_ranges"])
        
        wb.close()
        
        # VBA extraction
        if HAS_OLETOOLS and path.suffix.lower() == '.xlsm':
            try:
                vba_parser = olevba.VBA_Parser(str(path))
                if vba_parser.detect_vba_macros():
                    analysis["summary"]["has_vba"] = True
                    for (_, _, vba_filename, vba_code) in vba_parser.extract_macros():
                        analysis["vba_modules"][vba_filename] = vba_code
                vba_parser.close()
            except Exception as e:
                analysis["vba_error"] = str(e)
        
        return [TextContent(type="text", text=json.dumps(analysis, indent=2))]
    
    # -------------------------------------------------------------------------
    # UNKNOWN TOOL
    # -------------------------------------------------------------------------
    else:
        return [TextContent(type="text", text=f"Unknown tool: {name}")]


# =============================================================================
# MAIN
# =============================================================================

async def main():
    """Run the MCP server."""
    async with stdio_server() as (read_stream, write_stream):
        await server.run(
            read_stream,
            write_stream,
            server.create_initialization_options()
        )


if __name__ == "__main__":
    import asyncio
    asyncio.run(main())
